from django.http import HttpRequest, HttpResponse, HttpResponseRedirect,FileResponse, JsonResponse, Http404
from django.core.paginator import Paginator
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from django.contrib.auth import REDIRECT_FIELD_NAME
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.views import View
from django.views.generic import TemplateView
from django.urls import reverse, reverse_lazy
from .models import *
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import Group
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission
from django.db import connections
from django.db.models import Q
from collections import namedtuple
from django import forms
from datetime import time, datetime, timedelta,date
from django.contrib.auth.views import redirect_to_login
from django.contrib.auth.decorators import permission_required
from django.db.models import Count
from django.conf import settings
from wsgiref.util import FileWrapper
from docxtpl import DocxTemplate
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment
from openpyxl.utils import get_column_letter
from django.db import IntegrityError, transaction
import holidays
from workalendar.europe import Italy
from openpyxl.writer.excel import save_virtual_workbook
from cedolini import settaggio_ore as so
import io
import os
import json

@csrf_exempt
@login_required
def generaExcelDipendenti(request):
    date = datetime.now().date()
    try:
        if request.method == "GET" and request.GET.get('dipendenti'):
            data = request.GET.get('dipendenti')
            query = AnaDipendenti.objects.filter(stato='Attivo').values_list("nome","cognome","sede__nome_sede","area__nome_area","societa__nome_societa","tipo_contratto__nome_contratto","mansione__tipo_mansione").order_by('cognome')
        if request.method == "GET" and request.GET.get('area'):
            data = request.GET.get('area')
            query = AnaDipendenti.objects.filter(area=data,stato='Attivo').values_list("nome","cognome","sede__nome_sede","area__nome_area","societa__nome_societa","tipo_contratto__nome_contratto","mansione__tipo_mansione").order_by('cognome')
        if request.method == "GET" and request.GET.get('sede'):
            data = request.GET.get('sede')
            query = AnaDipendenti.objects.filter(sede=data,stato='Attivo').values_list("nome","cognome","sede__nome_sede","area__nome_area","societa__nome_societa","tipo_contratto__nome_contratto","mansione__tipo_mansione").order_by('cognome')
        if request.method == "GET" and request.GET.get('societa'):
            data = request.GET.get('societa')
            query = AnaDipendenti.objects.filter(societa=data,stato='Attivo').values_list("nome","cognome","sede__nome_sede","area__nome_area","societa__nome_societa","tipo_contratto__nome_contratto","mansione__tipo_mansione").order_by('cognome')
        if request.method == "GET" and request.GET.get('contratto'):
            data = request.GET.get('contratto')
            query = AnaDipendenti.objects.filter(tipo_contratto=data,stato='Attivo').values_list("nome","cognome","sede__nome_sede","area__nome_area","societa__nome_societa","tipo_contratto__nome_contratto","mansione__tipo_mansione").order_by('cognome')
        wb = Workbook()
    
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 30
            cursor.value = "Dipendente"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Sede"
            
            cursor = sheet.cell(row=1,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 15
            cursor.value = "Area"
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 20
            cursor.value = "Società"
            
            cursor = sheet.cell(row=1,column=5)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['E'].width = 25
            cursor.value = "Contratto"
            
            cursor = sheet.cell(row=1,column=5)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['F'].width = 30
            cursor.value = "Mansione"
            
            for index,values in enumerate(query, start=1):
                
                cursor = sheet.cell(row=1+index,column=1)
                cursor.value = f'{values[1]} {values[0]}'
                
                cursor = sheet.cell(row=1+index,column=2)
                cursor.value = values[2]
                
                cursor = sheet.cell(row=1+index,column=3)
                cursor.value = values[3]
                
                cursor = sheet.cell(row=1+index,column=4)
                cursor.value = values[4]
                
                cursor = sheet.cell(row=1+index,column=5)
                cursor.value = values[5]
                
                cursor = sheet.cell(row=1+index,column=5)
                cursor.value = values[6]

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Lista Dipendenti-{str(data)}-date.xlsx' 
            
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
   

@csrf_exempt
@login_required
def generaExcelMansioni(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('mansioni'):
            query = request.GET.get('mansioni') or "" 
            query = Mansione.objects.filter(Q(tipo_mansione__icontains=query)).order_by('tipo_mansione')
        elif request.method == "GET":
            query = Mansione.objects.all().order_by('tipo_mansione')
        wb = Workbook()
    
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 100
            cursor.value = "Mansione"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Note"

            for index,values in enumerate(query, start=1):
                cursor = sheet.cell(row=1+index,column=1)
                cursor.value = values.tipo_mansione
                
                cursor = sheet.cell(row=1+index,column=2)
                cursor.value = values.note


            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Lista Mansioni al {str(date)}.xlsx' 
            
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
   

@csrf_exempt
@login_required
def generaExcelContratti(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('contratti'):
            query = request.GET.get('contratti') or ""
            query = TipoContratto.objects.filter(Q(nome_contratto__icontains=query)).order_by('nome_contratto')
        elif request.method == "GET":
            query = TipoContratto.objects.all().order_by('nome_contratto')
        wb = Workbook()
    
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 100
            cursor.value = "Tipo Contratto"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Note"

            for index,values in enumerate(query, start=1):
                cursor = sheet.cell(row=1+index,column=1)
                cursor.value = values.nome_contratto
                
                cursor = sheet.cell(row=1+index,column=2)
                cursor.value = values.note


            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Tipologie Contratti al {str(date)}.xlsx' 
            
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
   
   
@csrf_exempt
@login_required
def generaExcelIstruzione(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('istruzioni'):
            query = request.GET.get('istruzioni') or ""
            query = Istruzione.objects.filter(Q(tipo_istruzione__icontains=query)).order_by('tipo_istruzione')
        elif request.method == "GET":
            query = Istruzione.objects.all().order_by('tipo_istruzione')
        wb = Workbook()
    
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 100
            cursor.value = "Tipo Istruzione"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Note"

            for index,values in enumerate(query, start=1):
                cursor = sheet.cell(row=1+index,column=1)
                cursor.value = values.tipo_istruzione
                
                cursor = sheet.cell(row=1+index,column=2)
                cursor.value = values.note


            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Tipo Istruzione al {str(date)}.xlsx' 
            
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
   

@csrf_exempt
@login_required
def generaExcelSocieta(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('societa'):
            query = request.GET.get('societa') or ""
            query = Societa.objects.filter(Q(nome_societa__icontains=query)).order_by('nome_societa')
        elif request.method == "GET":
            query = Societa.objects.all().order_by('nome_societa')
        wb = Workbook()
    
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 100
            cursor.value = "Società"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Note"

            for index,values in enumerate(query, start=1):
                cursor = sheet.cell(row=1+index,column=1)
                cursor.value = values.nome_societa
                
                cursor = sheet.cell(row=1+index,column=2)
                cursor.value = values.note


            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Società al {str(date)}.xlsx' 
            
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
   
   
@csrf_exempt
@login_required
def generaExcelSedi(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('sedi'):
            query = request.GET.get('sedi') or ""
            query = Sede.objects.filter(Q(nome_sede__icontains=query)).order_by('nome_sede')
        elif request.method == "GET":
            query = Sede.objects.all().order_by('nome_sede')
        wb = Workbook()
    
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 100
            cursor.value = "Sedi"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Note"

            for index,values in enumerate(query, start=1):
                cursor = sheet.cell(row=1+index,column=1)
                cursor.value = values.nome_sede
                
                cursor = sheet.cell(row=1+index,column=2)
                cursor.value = values.note


            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Sedi al {str(date)}.xlsx' 
            
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
 
@csrf_exempt
@login_required
def generaExcelArea(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('area'):
            query = request.GET.get('area') or ""
            query = Area.objects.filter(Q(nome_area__icontains=query)).order_by('nome_area')
        elif request.method == "GET":
            query = Area.objects.all().order_by('nome_area')
        wb = Workbook()
    
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 100
            cursor.value = "Aree"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Note"

            for index,values in enumerate(query, start=1):
                cursor = sheet.cell(row=1+index,column=1)
                cursor.value = values.nome_area
                
                cursor = sheet.cell(row=1+index,column=2)
                cursor.value = values.note


            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Aree al {str(date)}.xlsx' 
            
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    


@csrf_exempt
@login_required
def generaExcelCapiArea(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('capiarea'):
            query = request.GET.get('area') or ""
            query = CapoArea.objects.filter(Q(nomecompleto__icontains=query)).order_by('nomecompleto')
        elif request.method == "GET":
            query = CapoArea.objects.all().order_by('nomecompleto')
        wb = Workbook()
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 25
            cursor.value = "Capo Area"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 30
            cursor.value = "Sede"
            
            cursor = sheet.cell(row=1,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 30
            cursor.value = "Area"
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 30
            cursor.value = "Societa"


            for index,values in enumerate(query, start=1):
                try:                    
                    cursor = sheet.cell(row=1+index,column=1)
                    cursor.value = values.nomecompleto
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=2)
                    cursor.value = str(values.id_dipendente.sede).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=3)
                    cursor.value = str(values.id_dipendente.area).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=4)
                    cursor.value = str(values.id_dipendente.societa).title()
                except Exception as err:
                    print(err)

            
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="Capi Area - etc al {str(date)}.xlsx' 
           
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    

@csrf_exempt
@login_required
def generaExcelResponsabili(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('responsabili'):
            query = request.GET.get('area') or ""
            query = Responsabili.objects.filter(Q(nomecompleto__icontains=query)).order_by('nomecompleto')
        elif request.method == "GET":
            query = Responsabili.objects.all().order_by('nomecompleto')
        wb = Workbook()
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 25
            cursor.value = "Responsabile"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 30
            cursor.value = "Sede"
            
            cursor = sheet.cell(row=1,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 30
            cursor.value = "Area"
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 30
            cursor.value = "Societa"


            for index,values in enumerate(query, start=1):
                try:                    
                    cursor = sheet.cell(row=1+index,column=1)
                    cursor.value = values.nomecompleto
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=2)
                    cursor.value = str(values.id_dipendente.sede).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=3)
                    cursor.value = str(values.id_dipendente.area).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=4)
                    cursor.value = str(values.id_dipendente.societa).title()
                except Exception as err:
                    print(err)

            
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="Responsabili - etc al {str(date)}.xlsx' 
           
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    


@csrf_exempt
@login_required
def generaExcelResponsabiliSede(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('responsabili-sede'):
            query = request.GET.get('area') or ""
            query = ResponsabiliSede.objects.filter(Q(nomecompleto__icontains=query)).order_by('nomecompleto')
        elif request.method == "GET":
            query = ResponsabiliSede.objects.all().order_by('nomecompleto')
        wb = Workbook()
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 25
            cursor.value = "Responsabile"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 30
            cursor.value = "Sede"
            
            cursor = sheet.cell(row=1,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 30
            cursor.value = "Area"
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 30
            cursor.value = "Societa"


            for index,values in enumerate(query, start=1):
                try:                    
                    cursor = sheet.cell(row=1+index,column=1)
                    cursor.value = values.nomecompleto
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=2)
                    cursor.value = str(values.id_dipendente.sede).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=3)
                    cursor.value = str(values.id_dipendente.area).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=4)
                    cursor.value = str(values.id_dipendente.societa).title()
                except Exception as err:
                    print(err)

            
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="Responsabili - etc al {str(date)}.xlsx' 
           
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    

@csrf_exempt
@login_required
def generaExcelDirigenti(request):
    date = datetime.now().date()
    query = ""
    try:
        if request.method == "GET" and request.GET.get('dirigenti'):
            query = request.GET.get('area') or ""
            query = Dirigenti.objects.filter(Q(nomecompleto__icontains=query)).order_by('nomecompleto')
        elif request.method == "GET":
            query = Dirigenti.objects.all().order_by('nomecompleto')
        wb = Workbook()
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 25
            cursor.value = "Responsabile Sede"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 30
            cursor.value = "Sede"
            
            cursor = sheet.cell(row=1,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 30
            cursor.value = "Area"
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 30
            cursor.value = "Societa"


            for index,values in enumerate(query, start=1):
                try:                    
                    cursor = sheet.cell(row=1+index,column=1)
                    cursor.value = values.nomecompleto
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=2)
                    cursor.value = str(values.id_dipendente.sede).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=3)
                    cursor.value = str(values.id_dipendente.area).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=4)
                    cursor.value = str(values.id_dipendente.societa).title()
                except Exception as err:
                    print(err)

            
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="Dirigenti - etc al {str(date)}.xlsx' 
           
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    


@login_required
@permission_required("gestione.change_richiesteaccettate")
@csrf_exempt
def scaricaRichiestePersonale(request,data,dipendente=None):
    wb = Workbook()
    today = datetime.now().date()
    try:
        if dipendente == "none":
            queryset = RichiesteAccettate.objects.filter(in_corso=0).order_by("-data_creazione")
            in_permesso = queryset.filter(data_fine_permesso__gte=today).count()
            rientranti = queryset.filter(data_fine_permesso__gte=today).count()
            sheet = wb.create_sheet(f'Richieste - {str(data)}')
            del wb['Sheet']
            cell_range = 'A1:I1'
            cursor = sheet.merge_cells(cell_range)
            cursor = sheet.cell(row=1,column=1)           
            cursor.alignment = Alignment(horizontal='center')

            cursor.value = f"Richieste aggiornate alla data {str(data)}"

            cursor = sheet.cell(row=2,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 25
            cursor.value = "Dipendenti"
            cursor = sheet.cell(row=2,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 25
            cursor.value = "Area"
            cursor = sheet.cell(row=2,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 15
            cursor.value = "Sede"
            cursor = sheet.cell(row=2,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 25
            cursor.value = "Responsabile"
            cursor = sheet.cell(row=2,column=5)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['E'].width = 16
            cursor.value = "Tipo Permesso"
            cursor = sheet.cell(row=2,column=6)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['F'].width = 10
            cursor.value = "Inizio"
            cursor = sheet.cell(row=2,column=7)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['G'].width = 10
            cursor.value = "Fine"
            cursor = sheet.cell(row=2,column=8)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['H'].width = 10
            cursor.value = "Orario"
            cursor = sheet.cell(row=2,column=9)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['I'].width = 10
            cursor.value = "Urgente?"
            
            for index, el in enumerate(queryset):
                try:                    
                    cursor = sheet.cell(row=3+index,column=1)
                    cursor.value = f'{(el.id_richieste.nominativo).title()}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=2)
                    cursor.value = f'{(el.id_richieste.id_dipendente_richiesta.area)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=3)
                    cursor.value = f'{(el.id_richieste.id_dipendente_richiesta.sede)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=4)
                    cursor.value = f'{el.id_capoarea_richieste.nomecompleto}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=5)
                    if el.id_richieste.id_permessi_richieste != None: 
                        cursor.value = f'{el.id_richieste.id_permessi_richieste.codicepermesso}'
                    else: cursor.value = "Permesso Orario"                
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=6)
                    cursor.value = f'{(el.data_inizio_permesso)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=7)
                    cursor.value = f'{(el.data_fine_permesso)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=8)
                    if el.ora_inizio_permesso:
                        cursor.value = f'{(el.ora_inizio_permesso) - {el.ora_fine_permesso}}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=9)
                    if el.id_richieste.urgente:
                        cursor.value = "Si"
                    else:
                        cursor.value = "No"              
                except Exception as err:
                    print(err)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Richieste {str(data)}.xlsx'
            return response
        else:
            queryset = RichiesteAccettate.objects.filter(Q(id_richieste__id_dipendente_richiesta__cognome__icontains=dipendente)|Q(id_richieste__id_dipendente_richiesta__nome__icontains=dipendente)).filter(in_corso=0).order_by("-data_creazione")
            in_permesso = queryset.filter(data_fine_permesso__gte=today).count()
            rientranti = queryset.filter(data_fine_permesso__gte=today).count()
            sheet = wb.create_sheet(f'Richieste - {str(data)}')
            del wb['Sheet']
            cell_range = 'A1:I1'
            sheet.merge_cells(cell_range) 
            cursor = sheet.cell(row=1,column=1)           
            cursor.alignment = Alignment(horizontal='center')
            
            if dipendente != None:
                cursor.value = f'Richieste aggiornate alla data {str(data)}, filtro utilizzato: "{dipendente}"'
            else:
                cursor.value = f'Richieste aggiornate alla data {str(data)}'

            cursor = sheet.cell(row=2,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 25
            cursor.value = "Dipendenti"
            cursor = sheet.cell(row=2,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 25
            cursor.value = "Area"
            cursor = sheet.cell(row=2,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 15
            cursor.value = "Sede"
            cursor = sheet.cell(row=2,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 25
            cursor.value = "Responsabile"
            cursor = sheet.cell(row=2,column=5)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['E'].width = 16
            cursor.value = "Tipo Permesso"
            cursor = sheet.cell(row=2,column=6)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['F'].width = 10
            cursor.value = "Inizio"
            cursor = sheet.cell(row=2,column=7)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['G'].width = 10
            cursor.value = "Fine"
            cursor = sheet.cell(row=2,column=8)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['H'].width = 10
            cursor.value = "Orario"
            cursor = sheet.cell(row=2,column=9)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['I'].width = 10
            cursor.value = "Urgente?"

            for index, el in enumerate(queryset):
                try:                    
                    cursor = sheet.cell(row=3+index,column=1)
                    cursor.value = f'{(el.id_richieste.nominativo).title()}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=2)
                    cursor.value = f'{(el.id_richieste.id_dipendente_richiesta.area)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=3)
                    cursor.value = f'{(el.id_richieste.id_dipendente_richiesta.sede)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=4)
                    cursor.value = f'{el.id_capoarea_richieste.nomecompleto}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=5)
                    if el.id_richieste.id_permessi_richieste != None: 
                        cursor.value = f'{el.id_richieste.id_permessi_richieste.codicepermesso}'
                    else: cursor.value = "Permesso Orario"                 
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=6)
                    cursor.value = f'{(el.data_inizio_permesso)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=7)
                    cursor.value = f'{(el.data_fine_permesso)}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=8)
                    if el.ora_inizio_permesso:
                        cursor.value = f'{(el.ora_inizio_permesso) - {el.ora_fine_permesso}}'
                except Exception as err:
                    print(err)
                try:                    
                    cursor = sheet.cell(row=3+index,column=9)
                    if el.id_richieste.urgente:
                        cursor.value = "Si"
                    else:
                        cursor.value = "No"              
                except Exception as err:
                    print(err)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Richieste - {str(data)}.xlsx'
            return response
    except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  

@csrf_exempt
@login_required
@permission_required("gestione.change_anadipendenti")
def generaContratti(request,dipendente):
    for el in request:
        print(el)
    element=request.GET.get('dipendente')
    query = ""
    try:
        if request.method == "GET" and request.GET.get('dipendente'):
            query = request.GET.get('dipendente') or ""
            query = Contratti.objects.filter(Q(id_dip__cognome__icontains=query)|Q(id_dip__nome__icontains=query)).order_by('id_dip__cognome')
        elif request.method == "GET":
            query = Contratti.objects.all().order_by('id_dip__cognome')
        wb = Workbook()
        del wb['Sheet']
        cell_range = 'A1:G1'
        sheet.merge_cells(cell_range) 
        cursor = sheet.cell(row=1,column=1)           
        cursor.alignment = Alignment(horizontal='center')
        cursor.value = f"Contratti"
        
        try:
            sheet = wb.create_sheet(f'SheetOne')
            del wb['Sheet']
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 25
            cursor.value = "Dipendente"
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 30
            cursor.value = "Società"
            
            cursor = sheet.cell(row=1,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 30
            cursor.value = "Tipologia"
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 30
            cursor.value = "CCNL"
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['E'].width = 15
            cursor.value = "Percentuale"

            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['F'].width = 15
            cursor.value = "Data Inizio"

            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['G'].width = 15
            cursor.value = "Data Fine"


            for index,values in enumerate(query, start=1):
                try:                    
                    cursor = sheet.cell(row=1+index,column=1)
                    cursor.value = values.id_dip.nominativo
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=2)
                    cursor.value = str(values.id_societa.nome_societa).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=3)
                    cursor.value = str(values.tipologia.nome_contratto).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=4)
                    cursor.value = str(values.ccnl).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=5)
                    cursor.value = str(values.percentuale).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=6)
                    cursor.value = str(values.datainizio).title()
                except Exception as err:
                    print(err)
                try:
                    cursor = sheet.cell(row=1+index,column=7)
                    cursor.value = str(values.datafine).title()
                except Exception as err:
                    print(err)

            
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="Contratti - al {str(date)}.xlsx' 
           
            return response
        except:
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))  
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    