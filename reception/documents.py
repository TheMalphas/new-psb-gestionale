from django.http import HttpRequest, HttpResponse, HttpResponseRedirect,FileResponse, JsonResponse, Http404
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Sum
from django.contrib.auth import REDIRECT_FIELD_NAME
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.views import View
from django.views.generic import TemplateView
from django.urls import reverse, reverse_lazy
from .models import Richieste, AuthUser, AnaDipendenti, BancaOrari, CapoArea, Permessi, RichiesteAccettate, Ingressidip, Ritardi_Anticipi, Area
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import Group
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission
from django.db import connections
from collections import namedtuple
from django import forms
from .forms import UpdateEntrata, UpdateUscita, UpdateDoppiaEntrata, UpdateDoppiaUscita, CreaEntrata, CreaEntrataDoppia, EntrateForm, UsciteForm
from datetime import time, datetime, timedelta,date
from django.contrib.auth.views import redirect_to_login
from django.utils.translation import get_language, activate
from django.db.models import Count
from django.conf import settings
from wsgiref.util import FileWrapper
from docxtpl import DocxTemplate
from io import BytesIO
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment
from openpyxl.utils import get_column_letter
from django.db import IntegrityError, transaction
import holidays
from workalendar.europe import Italy
from cedolini import settaggio_ore as so
import traceback
import io
import os
import json

def Gennaio(year=datetime.now().year):
    return "Gennaio",1,31

def Febbraio(year=datetime.now().year):
    if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
        return "Febbraio",2,29
    else:
        return "Febbraio",2,28

def Marzo(year=datetime.now().year):
    return "Marzo",3,31

def Aprile(year=datetime.now().year):
    return "Aprile",4,30

def Maggio(year=datetime.now().year):
    return "Maggio",5,31

def Giugno(year=datetime.now().year):
    return "Giugno",6,30

def Luglio(year=datetime.now().year):
    return "Luglio",7,31

def Agosto(year=datetime.now().year):
    return "Agosto",8,31

def Settembre(year=datetime.now().year):
    return "Settembre",9,30

def Ottobre(year=datetime.now().year):
    return "Ottobre",10,31

def Novembre(year=datetime.now().year):
    return "Novembre",11,30

def Dicembre(year=datetime.now().year):
    return "Dicembre",1,31

switcher = {
1: Gennaio,
2: Febbraio,
3: Marzo,
4: Aprile,
5: Maggio,
6: Giugno,
7: Luglio,
8: Agosto,
9: Settembre,
10: Ottobre,
11: Novembre,
12: Dicembre,
}

def switch(counter):
    return switcher.get(counter)()


def settimanainteressata(data):
    anno = str(data)[0:4]
    bisestile = ""
    if (int(anno)%4 == 0 and int(anno)%100 != 0) or (int(anno)%400 == 0) :
        bisestile = 1
    else :
        bisestile = 0
    
    mese = str(data)[-5:-3]
    mesi31 = ["01","03","05","06","08","10","12"]
    mesi30 = ["04","07","09","11"]
    
    if mese in mesi31:
        last_giorno = "31"
    elif mese in mesi30:
        last_giorno = "30"
    elif mese == "02" and bisestile == 1:
        last_giorno = "29"
    else: last_giorno = "28"
    
    datetime_object = datetime.strptime(f'{str(data)} 00:00:00', '%Y-%m-%d %H:%M:%S')
    day = datetime_object.weekday()
    daysOfTheWeek = [0,1,2,3,4,5,6]
    dayStart = None
    dayEnd = None
    
    if day == daysOfTheWeek[0]:
        dayStart = data 
        dayEnd = data + timedelta(days=6)
    elif day == daysOfTheWeek[1]:
        dayStart = data - timedelta(days=1)
        dayEnd = data + timedelta(days=5)
    elif day == daysOfTheWeek[2]:
        dayStart = data - timedelta(days=2)
        dayEnd = data + timedelta(days=4)
    elif day == daysOfTheWeek[3]:
        dayStart = data - timedelta(days=3)
        dayEnd = data + timedelta(days=3)
    elif day == daysOfTheWeek[4]:
        dayStart = data - timedelta(days=4)
        dayEnd = data + timedelta(days=2)
    elif day == daysOfTheWeek[5]:
        dayStart = data - timedelta(days=5)
        dayEnd = data + timedelta(days=1)
    elif day == daysOfTheWeek[6]:
        dayStart = data - timedelta(days=6)
        dayEnd = data

    data1 = dayStart
    data2 = dayEnd

    return data1, data2

def meseinteressato(data):
    data = str(data)
    anno = data[0:4]
    bisestile = ""
    if (int(anno)%4 == 0 and int(anno)%100 != 0) or (int(anno)%400 == 0) :
        bisestile = 1
    else :
        bisestile = 0
        
    mese = data[-5:-3]
    mesi31 = ["01","03","05","06","08","10","12"]
    mesi30 = ["04","07","09","11"]
    
    first_giorno = "01"
    last_giorno = ""
    
    if mese in mesi31:
        last_giorno = "31"
    elif mese in mesi30:
        last_giorno = "30"
    elif mese == "02" and bisestile == 1:
        last_giorno = "29"
    else: last_giorno = "28"
    
    data1 = f'{anno}-{mese}-{first_giorno}'
    data2 = f'{anno}-{mese}-{last_giorno}' 
    
    return data1, data2

def meseinteressato2(data):
    mese = str(data)
    anno = datetime.today().year
    
    bisestile = ""
    if (int(anno)%4 == 0 and int(anno)%100 != 0) or (int(anno)%400 == 0) :
        bisestile = 1
    else :
        bisestile = 0
        
    mesi31 = ["01","03","05","06","08","10","12"]
    mesi30 = ["04","07","09","11"]
    
    first_giorno = "01"
    last_giorno = ""
    
    if mese in mesi31:
        last_giorno = "31"
    elif mese in mesi30:
        last_giorno = "30"
    elif mese == "02" and bisestile == 1:
        last_giorno = "29"
    else: last_giorno = "28"
    
    data1 = f'{anno}-{mese}-{first_giorno}'
    data2 = f'{anno}-{mese}-{last_giorno}' 

    return data1, data2


#Funzione per calcolare l'anticipo

def ritardoFunc(value):
    arrivo = datetime.strptime('09:00:00','%H:%M:%S')
    uscita = datetime.strptime('18:00:00','%H:%M:%S')
    if value != None:
        orario = value.strftime('%H:%M:%S')
        orario = datetime.strptime(f'1900-01-01 {orario}', '%Y-%m-%d %H:%M:%S')
        
        if orario > arrivo:
            data = orario - arrivo
            return f'{data}'


#Funzione per calcolare il ritardo

def anticipoFunc(value):
    arrivo = datetime.strptime('09:00:00','%H:%M:%S')
    uscita = datetime.strptime('18:00:00','%H:%M:%S')
    if value != None:
        orario = value.strftime('%H:%M:%S')
        orario = datetime.strptime(f'1900-01-01 {orario}', '%Y-%m-%d %H:%M:%S')
        
        if orario < arrivo:
            data = arrivo - orario
            return f'{data}'

#Funzione per calcolare l'extra

def extraFunc(value):
    uscita = datetime.strptime('18:00:00','%H:%M:%S')
    if value != None:
        orario = value.strftime('%H:%M:%S')
        orario = datetime.strptime(f'1900-01-01 {orario}', '%Y-%m-%d %H:%M:%S')
        
        if orario > uscita:
            data = orario - uscita
            return f'{data}'


def calcoloOre(totOLav):
    state = 0
    for element in totOLav:
        if element[0] != None and element[1] != None:
            ele0 = element[0].strftime('%H:%M:%S')
            ele1 = element[1].strftime('%H:%M:%S')
            arrivo = datetime.strptime(f'1900-01-01 {ele0}','%Y-%m-%d %H:%M:%S')
            uscita = datetime.strptime(f'1900-01-01 {ele1}','%Y-%m-%d %H:%M:%S')
            diff = uscita-arrivo
            state += diff.total_seconds()
    lista=timedelta(seconds=state)
    return lista


def transDate(date):
    value = ((datetime.strptime(f'1900-01-01 {date}','%Y-%m-%d %H:%M:%S')))
    return value


@login_required
def generaReportExcelMassa(request):
    listaMesi = [1,2,3,4,5,6,7,8,9,10,11,12]
    if request.method == "POST" and request.POST.get("mese") and request.POST.get("mese") != None or request.POST.get("mese") != "":
        try:
            mese = int(request.POST.get("mese"))
            if mese in listaMesi:
                if mese < 10:
                    mese = f'0{str(mese)}'
                else: mese = str(mese)
                data = f'2023-{mese}-01'
                ingressi = Ingressidip.objects.filter(giorno=data,id_dip_ing__stato="Attivo").values_list("nominativo","entrata","uscita","tipo","id_dip_ing__societa__nome_societa","id_dip_ing__sede__nome_sede","id_dip_ing__area__nome_area").order_by('nominativo')
                wb = Workbook()
                giorni =  meseinteressato(data)
                anno = giorni[0][0:4]
                giorniFestivi = so.getHolidays(int(anno))
                festivita = []
                for el in giorniFestivi:
                    fes = f'{el.year}-'
                    if el.month < 10:
                        fes += f'0{el.month}-'
                        if el.day < 10:
                            fes += f'0{el.day}'
                        else: 
                            fes += f'{el.day}'
                    else:
                        fes += f'{el.month}-'
                        if el.day < 10:
                            fes += f'0{el.day}'
                        else: 
                            fes += f'{el.day}'
                    festivita.append(fes)
                
                for index,values in enumerate(ingressi, start=1):
                    sheet = wb.create_sheet(f'T_{values[0]}') 
                    cursor = sheet.cell(row=1,column=1)
                    cursor.font = Font(bold=True)
                    sheet.column_dimensions['A'].width = 30
                    cursor.value = "Dipendente"
                    
                    cursor = sheet.cell(row=1+1,column=1)
                    cursor.value = values[0]
                    
                    cursor = sheet.cell(row=1,column=2)
                    cursor.font = Font(bold=True)
                    sheet.column_dimensions['B'].width = 15
                    cursor.value = "Società"
                    
                    cursor = sheet.cell(row=1+1,column=2)
                    cursor.value = values[4]
                    
                    cursor = sheet.cell(row=1,column=3)
                    cursor.font = Font(bold=True)
                    sheet.column_dimensions['C'].width = 15
                    cursor.value = "Sede"
                    
                    cursor = sheet.cell(row=1+1,column=3)
                    cursor.value = values[5]
                    
                    cursor = sheet.cell(row=1,column=4)
                    cursor.font = Font(bold=True)
                    sheet.column_dimensions['D'].width = 20
                    cursor.value = "Area"
                    
                    cursor = sheet.cell(row=1+1,column=4)
                    cursor.value = values[6]
                    
                    cursor = sheet.cell(row=1+3,column=1)
                    cursor.font = Font(bold=True)
                    cursor.value = "Giorno"
                    
                    cursor = sheet.cell(row=1+3,column=2)
                    cursor.font = Font(bold=True)
                    cursor.value = "Permesso"
                    
                    cursor = sheet.cell(row=1+3,column=3)
                    cursor.font = Font(bold=True)
                    cursor.value = "Entrata"
                    
                    cursor = sheet.cell(row=1+3,column=4)
                    cursor.font = Font(bold=True)
                    cursor.value = "Uscita"
                    
                    cursor = sheet.cell(row=1+3,column=5)
                    cursor.font = Font(bold=True)
                    cursor.value = "Anticipo"
                    
                    cursor = sheet.cell(row=1+3,column=6)
                    cursor.font = Font(bold=True)
                    cursor.value = "Ritardo"
                    
                    cursor = sheet.cell(row=1+3,column=7)
                    cursor.font = Font(bold=True)
                    cursor.value = "Extra"
                    cursor = sheet.cell(row=1+3,column=8)
                    cursor.font = Font(bold=True)
                    
                    cursor.value = "Possibili Stra."
                    cursor = sheet.cell(row=1+3,column=10)
                    cursor.font = Font(bold=True)
                    sheet.column_dimensions['J'].width = 30
                    cursor.value = "Giorni Lavorati"
                    
                    cursor = sheet.cell(row=1+4,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Giorni di assenza"
                    
                    cursor = sheet.cell(row=1+5,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Giorni con mancata timbratura"
                    
                    cursor = sheet.cell(row=1+6,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Giorni con possibili straordinari"
                    
                    cursor = sheet.cell(row=1+7,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Ore Totali lavorate"
                    
                    cursor = sheet.cell(row=1+8,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Tot. Anticipo (Entrata)"
                    
                    cursor = sheet.cell(row=1+9,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Tot. Ritardo (Entrata)"
                    
                    cursor = sheet.cell(row=1+10,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Tot. Extra (Uscita)"
                    
                    cursor = sheet.cell(row=1+11,column=10)
                    cursor.font = Font(bold=True)
                    cursor.value = "Tot. Poss. Stra Mese"

                    query = Ingressidip.objects.filter(nominativo=values[0],giorno__range=[giorni[0],giorni[1]]).values_list("giorno","entrata","uscita","tipo","tipo_permesso__codicepermesso","id_permesso__da_ora_richiesta","id_permesso__a_ora_richiesta","in_permesso").order_by("giorno")
                    totDLav = []
                    totOLav = []
                    totELav = []
                    totNotLav = []
                    totNoBadge = []
                    totPossExtra = []
                    totAnt = 0
                    totRit = 0
                    totExtra = 0
                    totStraDay = 0
                    
                    for idx, element in enumerate(query,start=1):
                        extraDay = 0
                        antiDay = 0
                        StraDay = 0
                        cursor = sheet.cell(row=idx+4,column=1)
                        cursor.font = Font(bold=False)
                        cursor.alignment = Alignment(horizontal="right")
                        day = str(element[0])
                        itaDay = f'{day[8:10]}-{day[5:7]}-{day[0:4]}'
                        cursor.value = itaDay
                        
                        if element[4] != None and element[0].weekday() != 5 and element[0].weekday() != 6:
                            cursor = sheet.cell(row=idx+4,column=2)
                            cursor.alignment = Alignment(horizontal="right")
                            cursor.font = Font(bold=True)
                            cursor.value = f'{element[4]}'
                            dayOne = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                            dayTwo = ((datetime.strptime(f'1900-01-01 08:00:00','%Y-%m-%d %H:%M:%S')))
                            if datetime.today().now().date() > element[0]:
                                totOLav.append((dayOne.time(),dayTwo.time()))
                                if element[0].weekday() != 5 and element[0].weekday() != 6 and element[1] == None and element[2] == None and element[7] == 0:
                                    totNotLav.append(element[0])

                        elif element[5] != None and element[6] != None:
                            cursor = sheet.cell(row=idx+4,column=2)
                            cursor.alignment = Alignment(horizontal="right")
                            cursor.font = Font(bold=True)
                            cursor.value = f'{str(element[5])[0:5]} - {str(element[6])[0:5]}'
                        
                        if element[1] and (element[0].weekday() != 5 and element[0].weekday() != 6):
                            totDLav.append(element[0])
                        
                        if datetime.today().now().date() > element[0]:
                            if element[0] not in festivita:
                                if element[2] == None and (element[0].weekday() != 5 and element[0].weekday() != 6):
                                    totNoBadge.append(element[0])
                                elif element[1] == None and (element[0].weekday() != 5 and element[0].weekday() != 6):
                                    totNoBadge.append(element[0])
                            
                        cursor = sheet.cell(row=idx+4,column=3)
                        for el in festivita:
                            if el == day and not(element[1]):
                                cursor.alignment = Alignment(horizontal="right")
                                cursor.font = Font(bold=True,color='FF0000')
                                cursor.value = "Festa"
                                break
                            elif el == day and element[1]:
                                cursor.alignment = Alignment(horizontal="right")
                                cursor.font = Font(bold=True,color='FF0000')
                                cursor.value = element[1]
                                cursor2 = sheet.cell(row=idx+4,column=4)
                                cursor2.alignment = Alignment(horizontal="right")
                                cursor2.font = Font(bold=True,color='FF0000')
                                break
                            elif element[0].weekday() == 5 and not(element[1]):
                                cursor.alignment = Alignment(horizontal="right")
                                cursor.font = Font(bold=True,color='FF0000')
                                cursor.value = "Sabato"
                                break
                            elif element[0].weekday() == 5 and element[1]:
                                cursor.alignment = Alignment(horizontal="right")
                                cursor.font = Font(bold=True,color='FF0000')
                                cursor.value = element[1]
                                cursor2 = sheet.cell(row=idx+4,column=4)
                                cursor2.alignment = Alignment(horizontal="right")
                                cursor2.font = Font(bold=True,color='FF0000')
                                break
                            elif element[0].weekday() == 6 and not(element[1]):
                                cursor.alignment = Alignment(horizontal="right")
                                cursor.font = Font(bold=True,color='FF0000')
                                cursor.value = "Domenica"
                                break
                            elif element[0].weekday() == 6 and element[1]:
                                cursor.alignment = Alignment(horizontal="right")
                                cursor.font = Font(bold=True,color='FF0000')
                                cursor.value = element[1]
                                cursor2 = sheet.cell(row=idx+4,column=4)
                                cursor2.alignment = Alignment(horizontal="right")
                                cursor2.font = Font(bold=True,color='FF0000')
                                break
                            else:
                                cursor.value = element[1]
                    
                        cursor = sheet.cell(row=idx+4,column=4)
                        cursor.value = element[2]
                        
                        cursor = sheet.cell(row=idx+4,column=5)
                        cursor.alignment = Alignment(horizontal="right")
                        anti = anticipoFunc(element[1])

                        if anti and (element[0].weekday() != 5 or element[0].weekday() != 6):
                            antiPre = transDate(anti)
                            compare = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                            antiDaAgg = antiPre - compare
                            antiDay += antiDaAgg.total_seconds()
                            if antiDaAgg.total_seconds() > 1800:
                                StraDay += antiDaAgg.total_seconds()
                                totStraDay += antiDaAgg.total_seconds()
                                if element[0] not in totPossExtra:
                                    totPossExtra.append(element[0])
                            totAnt += antiDaAgg.total_seconds()
                            cursor.value = antiDaAgg
                        else: cursor.value = anti


                        cursor = sheet.cell(row=idx+4,column=6)
                        cursor.alignment = Alignment(horizontal="right")
                        
                        rita = ritardoFunc(element[1])
                        if rita and (element[0].weekday() != 5 or element[0].weekday() != 6):
                            ritaPre = transDate(rita)
                            compare2 = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                            ritaDaAgg = ritaPre - compare2
                            totRit += ritaDaAgg.total_seconds() 
                            totRit += 300 
                            ritaDaAgg = ritaDaAgg + timedelta(minutes = 5)
                            cursor.value = ritaDaAgg
                        else: cursor.value = rita
                        totOLav.append((element[1],element[2]))
                        
                        cursor = sheet.cell(row=idx+4,column=7)
                        cursor.alignment = Alignment(horizontal="right")
                        extra = extraFunc(element[2])
                        if extra and (element[0].weekday() != 5 or element[0].weekday() != 6):
                            extraPre = transDate(extra)
                            compare3 = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                            extraDaAgg = extraPre - compare3
                            if extraDaAgg.total_seconds() >= 1800:
                                StraDay += extraDaAgg.total_seconds()
                                totStraDay += extraDaAgg.total_seconds()
                                if element[0] not in totPossExtra:
                                    totPossExtra.append(element[0])
                            elif extraDaAgg.total_seconds() <= 1800:
                                totExtra += extraDaAgg.total_seconds()
                                cursor.value = extra
                        totELav.append((element[1],element[2]))
                        
                        cursor = sheet.cell(row=idx+4,column=8)
                        cursor.alignment = Alignment(horizontal="right")
                        if StraDay > 1800:
                            cursor.value = timedelta(seconds=StraDay)

                    cursor = sheet.cell(row=1+3,column=11)
                    cursor.value = len(totDLav)
                    
                    totali = calcoloOre(totOLav)
                    
                    cursor = sheet.cell(row=1+4,column=11)
                    cursor.value = len(totNotLav)
                    
                    cursor = sheet.cell(row=1+5,column=11)
                    cursor.value = len(totNoBadge)
                    
                    cursor = sheet.cell(row=1+6,column=11)
                    cursor.value = len(totPossExtra)
                    
                    cursor = sheet.cell(row=1+7,column=11)
                    cursor.value = totali
                    
                    cursor = sheet.cell(row=1+8,column=11)
                    cursor.value = timedelta(seconds=totAnt)
                    
                    cursor = sheet.cell(row=1+9,column=11)
                    cursor.value = timedelta(seconds=totRit)
                    
                    cursor = sheet.cell(row=1+10,column=11)
                    cursor.value = timedelta(seconds=totExtra)
                    
                    cursor = sheet.cell(row=1+11,column=11)
                    cursor.value = timedelta(seconds=totStraDay)

                try:
                    del wb['Sheet']
                except:
                    pass
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Ingressi-{str(data)}.xlsx'
                return response  
        except: 
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    

@login_required
def generaReportExcelMassaMeseCorrente(request):
    today = datetime.now().date()
    try:
        data = today
        ingressi = Ingressidip.objects.filter(giorno=data,id_dip_ing__stato="Attivo").values_list("nominativo","entrata","uscita","tipo","id_dip_ing__societa__nome_societa","id_dip_ing__sede__nome_sede","id_dip_ing__area__nome_area").order_by('nominativo')
        wb = Workbook()
        giorni =  meseinteressato(data)
        anno = giorni[0][0:4]
        giorniFestivi = so.getHolidays(int(anno))
        festivita = []
        for el in giorniFestivi:
            fes = f'{el.year}-'
            if el.month < 10:
                fes += f'0{el.month}-'
                if el.day < 10:
                    fes += f'0{el.day}'
                else: 
                    fes += f'{el.day}'
            else:
                fes += f'{el.month}-'
                if el.day < 10:
                    fes += f'0{el.day}'
                else: 
                    fes += f'{el.day}'
            festivita.append(fes)
        
        for index,values in enumerate(ingressi, start=1):
            sheet = wb.create_sheet(f'T_{values[0]}') 
            cursor = sheet.cell(row=1,column=1)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['A'].width = 30
            cursor.value = "Dipendente"
            
            cursor = sheet.cell(row=1+1,column=1)
            cursor.value = values[0]
            
            cursor = sheet.cell(row=1,column=2)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['B'].width = 15
            cursor.value = "Società"
            
            cursor = sheet.cell(row=1+1,column=2)
            cursor.value = values[4]
            
            cursor = sheet.cell(row=1,column=3)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['C'].width = 15
            cursor.value = "Sede"
            
            cursor = sheet.cell(row=1+1,column=3)
            cursor.value = values[5]
            
            cursor = sheet.cell(row=1,column=4)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['D'].width = 20
            cursor.value = "Area"
            
            cursor = sheet.cell(row=1+1,column=4)
            cursor.value = values[6]
            
            cursor = sheet.cell(row=1+3,column=1)
            cursor.font = Font(bold=True)
            cursor.value = "Giorno"
            
            cursor = sheet.cell(row=1+3,column=2)
            cursor.font = Font(bold=True)
            cursor.value = "Permesso"
            
            cursor = sheet.cell(row=1+3,column=3)
            cursor.font = Font(bold=True)
            cursor.value = "Entrata"
            
            cursor = sheet.cell(row=1+3,column=4)
            cursor.font = Font(bold=True)
            cursor.value = "Uscita"
            
            cursor = sheet.cell(row=1+3,column=5)
            cursor.font = Font(bold=True)
            cursor.value = "Anticipo"
            
            cursor = sheet.cell(row=1+3,column=6)
            cursor.font = Font(bold=True)
            cursor.value = "Ritardo"
            
            cursor = sheet.cell(row=1+3,column=7)
            cursor.font = Font(bold=True)
            cursor.value = "Extra"
            cursor = sheet.cell(row=1+3,column=8)
            cursor.font = Font(bold=True)
            
            cursor.value = "Possibili Stra."
            cursor = sheet.cell(row=1+3,column=10)
            cursor.font = Font(bold=True)
            sheet.column_dimensions['J'].width = 30
            cursor.value = "Giorni Lavorati"
            
            cursor = sheet.cell(row=1+4,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Giorni di assenza"
            
            cursor = sheet.cell(row=1+5,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Giorni con mancata timbratura"
            
            cursor = sheet.cell(row=1+6,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Giorni con possibili straordinari"
            
            cursor = sheet.cell(row=1+7,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Ore Totali lavorate"
            
            cursor = sheet.cell(row=1+8,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Tot. Anticipo (Entrata)"
            
            cursor = sheet.cell(row=1+9,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Tot. Ritardo (Entrata)"
            
            cursor = sheet.cell(row=1+10,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Tot. Extra (Uscita)"
            
            cursor = sheet.cell(row=1+11,column=10)
            cursor.font = Font(bold=True)
            cursor.value = "Tot. Poss. Stra Mese"

            query = Ingressidip.objects.filter(nominativo=values[0],giorno__range=[giorni[0],giorni[1]]).values_list("giorno","entrata","uscita","tipo","tipo_permesso__codicepermesso","id_permesso__da_ora_richiesta","id_permesso__a_ora_richiesta","in_permesso").order_by("giorno")
            totDLav = []
            totOLav = []
            totELav = []
            totNotLav = []
            totNoBadge = []
            totPossExtra = []
            totAnt = 0
            totRit = 0
            totExtra = 0
            totStraDay = 0
            
            for idx, element in enumerate(query,start=1):
                extraDay = 0
                antiDay = 0
                StraDay = 0
                cursor = sheet.cell(row=idx+4,column=1)
                cursor.font = Font(bold=False)
                cursor.alignment = Alignment(horizontal="right")
                day = str(element[0])
                itaDay = f'{day[8:10]}-{day[5:7]}-{day[0:4]}'
                cursor.value = itaDay
                
                if element[4] != None and element[0].weekday() != 5 and element[0].weekday() != 6:
                    cursor = sheet.cell(row=idx+4,column=2)
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True)
                    cursor.value = f'{element[4]}'
                    dayOne = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                    dayTwo = ((datetime.strptime(f'1900-01-01 08:00:00','%Y-%m-%d %H:%M:%S')))
                    if datetime.today().now().date() > element[0]:
                        totOLav.append((dayOne.time(),dayTwo.time()))
                        if element[0].weekday() != 5 and element[0].weekday() != 6 and element[1] == None and element[2] == None and element[7] == 0:
                            totNotLav.append(element[0])

                elif element[5] != None and element[6] != None:
                    cursor = sheet.cell(row=idx+4,column=2)
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True)
                    cursor.value = f'{str(element[5])[0:5]} - {str(element[6])[0:5]}'
                
                if element[1] and (element[0].weekday() != 5 and element[0].weekday() != 6):
                    totDLav.append(element[0])
                
                if datetime.today().now().date() > element[0]:
                    if element[0] not in festivita:
                        if element[2] == None and (element[0].weekday() != 5 and element[0].weekday() != 6):
                            totNoBadge.append(element[0])
                        elif element[1] == None and (element[0].weekday() != 5 and element[0].weekday() != 6):
                            totNoBadge.append(element[0])
                    
                cursor = sheet.cell(row=idx+4,column=3)
                for el in festivita:
                    if el == day and not(element[1]):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Festa"
                        break
                    elif el == day and element[1]:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = element[1]
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                    elif element[0].weekday() == 5 and not(element[1]):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Sabato"
                        break
                    elif element[0].weekday() == 5 and element[1]:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = element[1]
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                    elif element[0].weekday() == 6 and not(element[1]):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Domenica"
                        break
                    elif element[0].weekday() == 6 and element[1]:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = element[1]
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                    else:
                        cursor.value = element[1]
            
                cursor = sheet.cell(row=idx+4,column=4)
                cursor.value = element[2]
                
                cursor = sheet.cell(row=idx+4,column=5)
                cursor.alignment = Alignment(horizontal="right")
                anti = anticipoFunc(element[1])

                if anti and (element[0].weekday() != 5 or element[0].weekday() != 6):
                    antiPre = transDate(anti)
                    compare = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                    antiDaAgg = antiPre - compare
                    antiDay += antiDaAgg.total_seconds()
                    if antiDaAgg.total_seconds() > 1800:
                        StraDay += antiDaAgg.total_seconds()
                        totStraDay += antiDaAgg.total_seconds()
                        if element[0] not in totPossExtra:
                            totPossExtra.append(element[0])
                    totAnt += antiDaAgg.total_seconds()
                    cursor.value = antiDaAgg
                else: cursor.value = anti


                cursor = sheet.cell(row=idx+4,column=6)
                cursor.alignment = Alignment(horizontal="right")
                
                rita = ritardoFunc(element[1])
                if rita and (element[0].weekday() != 5 or element[0].weekday() != 6):
                    ritaPre = transDate(rita)
                    compare2 = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                    ritaDaAgg = ritaPre - compare2
                    totRit += ritaDaAgg.total_seconds() 
                    totRit += 300 
                    ritaDaAgg = ritaDaAgg + timedelta(minutes = 5)
                    cursor.value = ritaDaAgg
                else: cursor.value = rita
                totOLav.append((element[1],element[2]))
                
                cursor = sheet.cell(row=idx+4,column=7)
                cursor.alignment = Alignment(horizontal="right")
                extra = extraFunc(element[2])
                if extra and (element[0].weekday() != 5 or element[0].weekday() != 6):
                    extraPre = transDate(extra)
                    compare3 = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                    extraDaAgg = extraPre - compare3
                    if extraDaAgg.total_seconds() >= 1800:
                        StraDay += extraDaAgg.total_seconds()
                        totStraDay += extraDaAgg.total_seconds()
                        if element[0] not in totPossExtra:
                            totPossExtra.append(element[0])
                    elif extraDaAgg.total_seconds() <= 1800:
                        totExtra += extraDaAgg.total_seconds()
                        cursor.value = extra
                totELav.append((element[1],element[2]))
                
                cursor = sheet.cell(row=idx+4,column=8)
                cursor.alignment = Alignment(horizontal="right")
                if StraDay > 1800:
                    cursor.value = timedelta(seconds=StraDay)

            cursor = sheet.cell(row=1+3,column=11)
            cursor.value = len(totDLav)
            
            totali = calcoloOre(totOLav)
            
            cursor = sheet.cell(row=1+4,column=11)
            cursor.value = len(totNotLav)
            
            cursor = sheet.cell(row=1+5,column=11)
            cursor.value = len(totNoBadge)
            
            cursor = sheet.cell(row=1+6,column=11)
            cursor.value = len(totPossExtra)
            
            cursor = sheet.cell(row=1+7,column=11)
            cursor.value = totali
            
            cursor = sheet.cell(row=1+8,column=11)
            cursor.value = timedelta(seconds=totAnt)
            
            cursor = sheet.cell(row=1+9,column=11)
            cursor.value = timedelta(seconds=totRit)
            
            cursor = sheet.cell(row=1+10,column=11)
            cursor.value = timedelta(seconds=totExtra)
            
            cursor = sheet.cell(row=1+11,column=11)
            cursor.value = timedelta(seconds=totStraDay)

        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Ingressi-{str(data)}.xlsx'  

        return response
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   

@login_required
def generaReportExcelMassaSettimanaCorrente(request):
    data =datetime.now().date()
    ingressi = Ingressidip.objects.filter(giorno=data,id_dip_ing__stato="Attivo").values_list("nominativo","entrata","uscita","tipo","id_dip_ing__societa__nome_societa","id_dip_ing__sede__nome_sede","id_dip_ing__area__nome_area").order_by('nominativo')
    wb = Workbook()
    giorni = settimanainteressata(data)
    anno = str(giorni[0])[0:4]
    giorniFestivi = so.getHolidays(int(anno))
    festivita = []
    for el in giorniFestivi:
        fes = f'{el.year}-'
        if el.month < 10:
            fes += f'0{el.month}-'
            if el.day < 10:
                fes += f'0{el.day}'
            else: 
                fes += f'{el.day}'
        else:
            fes += f'{el.month}-'
            if el.day < 10:
                fes += f'0{el.day}'
            else: 
                fes += f'{el.day}'
        festivita.append(fes)
    
    for index,values in enumerate(ingressi, start=1):
        sheet = wb.create_sheet(f'T_{values[0]}') 
        cursor = sheet.cell(row=1,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = "Dipendente"
        
        cursor = sheet.cell(row=1+1,column=1)
        cursor.value = values[0]
        
        cursor = sheet.cell(row=1,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 15
        cursor.value = "Società"
        
        cursor = sheet.cell(row=1+1,column=2)
        cursor.value = values[4]
        
        cursor = sheet.cell(row=1,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 15
        cursor.value = "Sede"
        
        cursor = sheet.cell(row=1+1,column=3)
        cursor.value = values[5]
        
        cursor = sheet.cell(row=1,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 20
        cursor.value = "Area"
        
        cursor = sheet.cell(row=1+1,column=4)
        cursor.value = values[6]
        
        cursor = sheet.cell(row=1+3,column=1)
        cursor.font = Font(bold=True)
        cursor.value = "Giorno"
        
        cursor = sheet.cell(row=1+3,column=2)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=1+3,column=3)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=1+3,column=4)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=1+3,column=5)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=1+3,column=6)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=1+3,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Extra"
        cursor = sheet.cell(row=1+3,column=8)
        cursor.font = Font(bold=True)
        
        cursor.value = "Possibili Stra."
        cursor = sheet.cell(row=1+3,column=10)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['J'].width = 30
        cursor.value = "Giorni Lavorati"
        
        cursor = sheet.cell(row=1+4,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni di assenza"
        
        cursor = sheet.cell(row=1+5,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni con mancata timbratura"
        
        cursor = sheet.cell(row=1+6,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni con possibili straordinari"
        
        cursor = sheet.cell(row=1+7,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Ore Totali lavorate"
        
        cursor = sheet.cell(row=1+8,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Anticipo (Entrata)"
        
        cursor = sheet.cell(row=1+9,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Ritardo (Entrata)"
        
        cursor = sheet.cell(row=1+10,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Extra (Uscita)"
        
        cursor = sheet.cell(row=1+11,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Poss. Stra Mese"

        query = Ingressidip.objects.filter(nominativo=values[0],giorno__range=[giorni[0],giorni[1]]).values_list("giorno","entrata","uscita","tipo","tipo_permesso__codicepermesso","id_permesso__da_ora_richiesta","id_permesso__a_ora_richiesta","in_permesso").order_by("giorno")
        totDLav = []
        totOLav = []
        totELav = []
        totNotLav = []
        totNoBadge = []
        totPossExtra = []
        totAnt = 0
        totRit = 0
        totExtra = 0
        totStraDay = 0
        
        for idx, element in enumerate(query,start=1):
            extraDay = 0
            antiDay = 0
            StraDay = 0
            cursor = sheet.cell(row=idx+4,column=1)
            cursor.font = Font(bold=False)
            cursor.alignment = Alignment(horizontal="right")
            day = str(element[0])
            itaDay = f'{day[8:10]}-{day[5:7]}-{day[0:4]}'
            cursor.value = itaDay
            
            if element[4] != None and element[0].weekday() != 5 and element[0].weekday() != 6:
                cursor = sheet.cell(row=idx+4,column=2)
                cursor.alignment = Alignment(horizontal="right")
                cursor.font = Font(bold=True)
                cursor.value = f'{element[4]}'
                dayOne = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                dayTwo = ((datetime.strptime(f'1900-01-01 08:00:00','%Y-%m-%d %H:%M:%S')))
                if datetime.today().now().date() > element[0]:
                    totOLav.append((dayOne.time(),dayTwo.time()))
                    if element[0].weekday() != 5 and element[0].weekday() != 6 and element[1] == None and element[2] == None and element[7] == 0:
                        totNotLav.append(element[0])

            elif element[5] != None and element[6] != None:
                cursor = sheet.cell(row=idx+4,column=2)
                cursor.alignment = Alignment(horizontal="right")
                cursor.font = Font(bold=True)
                cursor.value = f'{str(element[5])[0:5]} - {str(element[6])[0:5]}'
            
            if element[1] and (element[0].weekday() != 5 and element[0].weekday() != 6):
                totDLav.append(element[0])
            
            if datetime.today().now().date() > element[0]:
                if element[0] not in festivita:
                    if element[2] == None and (element[0].weekday() != 5 and element[0].weekday() != 6):
                        totNoBadge.append(element[0])
                    elif element[1] == None and (element[0].weekday() != 5 and element[0].weekday() != 6):
                        totNoBadge.append(element[0])
                
            cursor = sheet.cell(row=idx+4,column=3)
            for el in festivita:
                if el == day and not(element[1]):
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True,color='FF0000')
                    cursor.value = "Festa"
                    break
                elif el == day and element[1]:
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True,color='FF0000')
                    cursor.value = element[1]
                    cursor2 = sheet.cell(row=idx+4,column=4)
                    cursor2.alignment = Alignment(horizontal="right")
                    cursor2.font = Font(bold=True,color='FF0000')
                    break
                elif element[0].weekday() == 5 and not(element[1]):
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True,color='FF0000')
                    cursor.value = "Sabato"
                    break
                elif element[0].weekday() == 5 and element[1]:
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True,color='FF0000')
                    cursor.value = element[1]
                    cursor2 = sheet.cell(row=idx+4,column=4)
                    cursor2.alignment = Alignment(horizontal="right")
                    cursor2.font = Font(bold=True,color='FF0000')
                    break
                elif element[0].weekday() == 6 and not(element[1]):
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True,color='FF0000')
                    cursor.value = "Domenica"
                    break
                elif element[0].weekday() == 6 and element[1]:
                    cursor.alignment = Alignment(horizontal="right")
                    cursor.font = Font(bold=True,color='FF0000')
                    cursor.value = element[1]
                    cursor2 = sheet.cell(row=idx+4,column=4)
                    cursor2.alignment = Alignment(horizontal="right")
                    cursor2.font = Font(bold=True,color='FF0000')
                    break
                else:
                    cursor.value = element[1]
        
            cursor = sheet.cell(row=idx+4,column=4)
            cursor.value = element[2]
            
            cursor = sheet.cell(row=idx+4,column=5)
            cursor.alignment = Alignment(horizontal="right")
            anti = anticipoFunc(element[1])

            if anti and (element[0].weekday() != 5 or element[0].weekday() != 6):
                antiPre = transDate(anti)
                compare = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                antiDaAgg = antiPre - compare
                antiDay += antiDaAgg.total_seconds()
                if antiDaAgg.total_seconds() > 1800:
                    StraDay += antiDaAgg.total_seconds()
                    totStraDay += antiDaAgg.total_seconds()
                    if element[0] not in totPossExtra:
                        totPossExtra.append(element[0])
                totAnt += antiDaAgg.total_seconds()
                cursor.value = antiDaAgg
            else: cursor.value = anti


            cursor = sheet.cell(row=idx+4,column=6)
            cursor.alignment = Alignment(horizontal="right")
            
            rita = ritardoFunc(element[1])
            if rita and (element[0].weekday() != 5 or element[0].weekday() != 6):
                ritaPre = transDate(rita)
                compare2 = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                ritaDaAgg = ritaPre - compare2
                totRit += ritaDaAgg.total_seconds() 
                totRit += 300 
                ritaDaAgg = ritaDaAgg + timedelta(minutes = 5)
                cursor.value = ritaDaAgg
            else: cursor.value = rita
            totOLav.append((element[1],element[2]))
            
            cursor = sheet.cell(row=idx+4,column=7)
            cursor.alignment = Alignment(horizontal="right")
            extra = extraFunc(element[2])
            if extra and (element[0].weekday() != 5 or element[0].weekday() != 6):
                extraPre = transDate(extra)
                compare3 = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                extraDaAgg = extraPre - compare3
                if extraDaAgg.total_seconds() >= 1800:
                    StraDay += extraDaAgg.total_seconds()
                    totStraDay += extraDaAgg.total_seconds()
                    if element[0] not in totPossExtra:
                        totPossExtra.append(element[0])
                elif extraDaAgg.total_seconds() <= 1800:
                    totExtra += extraDaAgg.total_seconds()
                    cursor.value = extra
            totELav.append((element[1],element[2]))
            
            cursor = sheet.cell(row=idx+4,column=8)
            cursor.alignment = Alignment(horizontal="right")
            if StraDay > 1800:
                cursor.value = timedelta(seconds=StraDay)

        cursor = sheet.cell(row=1+3,column=11)
        cursor.value = len(totDLav)
        
        totali = calcoloOre(totOLav)
        
        cursor = sheet.cell(row=1+4,column=11)
        cursor.value = len(totNotLav)
        
        cursor = sheet.cell(row=1+5,column=11)
        cursor.value = len(totNoBadge)
        
        cursor = sheet.cell(row=1+6,column=11)
        cursor.value = len(totPossExtra)
        
        cursor = sheet.cell(row=1+7,column=11)
        cursor.value = totali
        
        cursor = sheet.cell(row=1+8,column=11)
        cursor.value = timedelta(seconds=totAnt)
        
        cursor = sheet.cell(row=1+9,column=11)
        cursor.value = timedelta(seconds=totRit)
        
        cursor = sheet.cell(row=1+10,column=11)
        cursor.value = timedelta(seconds=totExtra)
        
        cursor = sheet.cell(row=1+11,column=11)
        cursor.value = timedelta(seconds=totStraDay)

    try:
        del wb['Sheet']
    except:
        pass
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Ingressi-{str(giorni[0])}-{str(giorni[1])}.xlsx'  

    return response

def getPeriodo(mese):
    listaMesi = ["Gennaio","Febbraio", "Marzo","Aprile","Maggio","Giugno","Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"]
    if mese in listaMesi:
        value = listaMesi.index(mese)
        value = value + 1
        if value < 10:
            valueNew = "0"
            valueNew += str(value)
            return valueNew
        else: return value


@login_required
def generaReportExcel(request,iddipendente=None):
    today = datetime.now() 
    anno = today.year
    user = request.user.pk

    try:
        iddipendente = request.POST.get('dipendente')
    
    except:
        pass
    
    if iddipendente != None:
        getId = AnaDipendenti.objects.get(id_dip=iddipendente)
        giorniFestivi = so.getHolidays(int(anno))
        festivita = []
        for el in giorniFestivi:
            fes = f'{el.year}-'
            if el.month < 10:
                fes += f'0{el.month}-'
                if el.day < 10:
                    fes += f'0{el.day}'
                else: 
                    fes += f'{el.day}'
            else:
                fes += f'{el.month}-'
                if el.day < 10:
                    fes += f'0{el.day}'
                else: 
                    fes += f'{el.day}'
            festivita.append(fes)
        ingressi = Ingressidip.objects.filter(id_dip_ing=getId.id_dip).values_list("nominativo","entrata","uscita","tipo","id_dip_ing__societa__nome_societa","id_dip_ing__sede__nome_sede","id_dip_ing__area__nome_area","id_dip_ing__tipo_contratto__nome_contratto").order_by('nominativo')
        data = datetime.now().month
        wb = Workbook()
        
        dataList = switch(data)
        sheet = wb.create_sheet(f'T_Ingressi {dataList[0]}')    
        cursor = sheet.cell(row=1,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = "Dipendente"
        
        cursor = sheet.cell(row=1+1,column=1)
        cursor.value = ingressi[0][0]
        
        cursor = sheet.cell(row=1,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 15
        cursor.value = "Società"
        
        cursor = sheet.cell(row=1+1,column=2)
        cursor.value = ingressi[0][4]
        
        cursor = sheet.cell(row=1,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 15
        cursor.value = "Sede"
        
        cursor = sheet.cell(row=1+1,column=3)
        cursor.value = ingressi[0][5]
        
        cursor = sheet.cell(row=1,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 20
        cursor.value = "Area"
        
        cursor = sheet.cell(row=1+1,column=4)
        cursor.value = ingressi[0][6]
        
        cursor = sheet.cell(row=1+3,column=1)
        cursor.font = Font(bold=True)
        cursor.value = "Giorno"
        
        cursor = sheet.cell(row=1+3,column=2)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=1+3,column=3)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=1+3,column=4)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=1+3,column=5)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['E'].width = 20
        cursor.value = "Seconda Entrata"
        
        cursor = sheet.cell(row=1+3,column=6)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['F'].width = 20
        cursor.value = "Seconda Uscita"
        
        cursor = sheet.cell(row=1+3,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=1+3,column=8)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=1+3,column=9)
        cursor.font = Font(bold=True)
        cursor.value = "Extra"
        
        sheet.column_dimensions['J'].width = 15
        cursor = sheet.cell(row=1+3,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Possibili Stra."
        
        cursor = sheet.cell(row=1+3,column=12)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['L'].width = 30
        cursor.value = "Giorni Lavorati"
        
        cursor = sheet.cell(row=1+4,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni di assenza"
        
        cursor = sheet.cell(row=1+5,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni con mancata timbratura"
        
        cursor = sheet.cell(row=1+6,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni con possibili straordinari"
        
        cursor = sheet.cell(row=1+7,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Ore Totali lavorate"
        
        cursor = sheet.cell(row=1+8,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Anticipo (Entrata)"
        
        cursor = sheet.cell(row=1+9,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Ritardo (Entrata)"
        
        cursor = sheet.cell(row=1+10,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Extra (Uscita)"
        
        cursor = sheet.cell(row=1+11,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Poss. Stra Mese"

        query = Ingressidip.objects.filter(id_dip_ing=getId.id_dip,giorno__range=[f'{str(today.year)}-{str(dataList[1])}-01',f'{str(anno)}-{str(dataList[1])}-{str(dataList[2])}']).values_list("giorno","entrata","uscita","tipo","tipo_permesso__codicepermesso","id_permesso__da_ora_richiesta","id_permesso__a_ora_richiesta","in_permesso","seconda_entrata","seconda_uscita").order_by("giorno")
        totDLav = []
        totOLav = []
        totELav = []
        totNotLav = []
        totNoBadge = []
        totPossExtra = []
        totAnt = 0
        totRit = 0
        totExtra = 0
        totStraDay = 0
        
        for idx, element in enumerate(query, start=1):
            extraDay = 0
            antiDay = 0
            StraDay = 0
            cursor = sheet.cell(row=idx+4,column=1)
            cursor.font = Font(bold=False)
            cursor.alignment = Alignment(horizontal="right")
            day = str(element[0])
            giornata = element[0]
            entrata = element[1]
            uscita = element[2]
            tipo = element[3]
            codice_permesso = element[4]
            inizio_richiesta = element[5]
            fine_richiesta = element[6]
            in_permesso = element[7]
            seconda_entrata = element[8]
            seconda_uscita = element[9]
            sabDom = giornata.weekday()
            itaDay = f'{day[8:10]}-{day[5:7]}-{day[0:4]}'
            cursor.value = itaDay
            
            if codice_permesso != None and sabDom != 5 and sabDom != 6:
                cursor = sheet.cell(row=idx+4,column=2)
                cursor.alignment = Alignment(horizontal="right")
                cursor.font = Font(bold=True)
                cursor.value = f'{codice_permesso}'
                dayOne = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                dayTwo = ((datetime.strptime(f'1900-01-01 08:00:00','%Y-%m-%d %H:%M:%S')))
                if datetime.today().now().date() > giornata:
                    totOLav.append((dayOne.time(),dayTwo.time()))
                    if sabDom != 5 and sabDom != 6 and entrata == None and uscita == None and in_permesso == 0 and seconda_entrata == None and seconda_uscita == None :
                        totNotLav.append(giornata)
            elif inizio_richiesta != None and fine_richiesta != None:
                cursor = sheet.cell(row=idx+4,column=2)
                cursor.alignment = Alignment(horizontal="right")
                cursor.font = Font(bold=True)
                cursor.value = f'{str(inizio_richiesta)[0:5]} - {str(fine_richiesta)[0:5]}'
            
            if entrata and (sabDom != 5 and sabDom != 6):
                totDLav.append(giornata)
            
            if datetime.today().now().date() > giornata:
                if giornata not in festivita:
                    if uscita == None and (sabDom != 5 and sabDom != 6):
                        totNoBadge.append(giornata)
                    elif entrata == None and (sabDom != 5 and sabDom != 6):
                        totNoBadge.append(giornata)
                
            cursor = sheet.cell(row=idx+4,column=3)
            for el in festivita:
                if el == day and not(entrata):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Festa"
                        break
                elif el == day and entrata:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = entrata
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                elif sabDom == 5 and not(entrata):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Sabato"
                        break
                elif sabDom == 5 and entrata:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = entrata
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                elif sabDom == 6 and not(entrata):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Domenica"
                        break
                elif sabDom == 6 and entrata:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = entrata
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                else:
                        cursor.value = entrata
                        
            
            cursor = sheet.cell(row=idx+4,column=4)
            cursor.value = uscita
            
            cursor = sheet.cell(row=idx+4,column=5)
            if seconda_entrata:
                cursor.value = seconda_entrata
            
            cursor = sheet.cell(row=idx+4,column=6)
            if seconda_uscita:
                cursor.value = seconda_uscita
            
            cursor = sheet.cell(row=idx+4,column=7)
            cursor.alignment = Alignment(horizontal="right")
            anti = anticipoFunc(entrata)
            if anti and (sabDom != 5 or sabDom != 6):
                antiPre = transDate(anti)
                compare = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                antiDaAgg = antiPre - compare
                antiDay += antiDaAgg.total_seconds()
                if antiDaAgg.total_seconds() > 1800:
                    StraDay += antiDaAgg.total_seconds()
                    totStraDay += antiDaAgg.total_seconds()
                    if giornata not in totPossExtra:
                        totPossExtra.append(giornata)
                totAnt += antiDaAgg.total_seconds()
                cursor.value = antiDaAgg
            else: cursor.value = anti

            cursor = sheet.cell(row=idx+4,column=8)
            cursor.alignment = Alignment(horizontal="right")
            
            rita = ritardoFunc(entrata)
            if rita and (sabDom != 5 or sabDom != 6):
                ritaPre = transDate(rita)
                compare2 = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                ritaDaAgg = ritaPre - compare2
                totRit += ritaDaAgg.total_seconds()
                totRit += 300 
                ritaDaAgg = ritaDaAgg + timedelta(minutes = 5)
                cursor.value = ritaDaAgg
            else: cursor.value = rita
            totOLav.append((entrata,uscita))
            
            secondaEntrata = seconda_entrata
            secondaUscita = seconda_uscita
            
            if secondaEntrata != None and secondaUscita != None:
                primaSeconda = transDate(secondaEntrata)
                comparePrima = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                primSecDaAgg = primaSeconda - comparePrima
                
                secondaSeconda = transDate(secondaUscita)
                compareSeconda = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                secoSecDaAgg = secondaSeconda - compareSeconda

                totPrim = primSecDaAgg.total_seconds()
                totSec = secoSecDaAgg.total_seconds()
                totSecBoth = totPrim + totSec
                totOLav.append([seconda_entrata,seconda_uscita])
            
            cursor = sheet.cell(row=idx+4,column=9)
            cursor.alignment = Alignment(horizontal="right")
            extra = extraFunc(uscita)
            if extra and (sabDom != 5 or sabDom != 6):
                extraPre = transDate(extra)
                compare3 = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                extraDaAgg = extraPre - compare3
                if extraDaAgg.total_seconds() >= 1800:
                    StraDay += extraDaAgg.total_seconds()
                    totStraDay += extraDaAgg.total_seconds()
                    if giornata not in totPossExtra:
                        totPossExtra.append(giornata)
                elif extraDaAgg.total_seconds() <= 1800:
                    totExtra += extraDaAgg.total_seconds()
                    cursor.value = extra
            totELav.append((entrata,uscita))
            
            cursor = sheet.cell(row=idx+4,column=10)
            cursor.alignment = Alignment(horizontal="right")
            if StraDay > 1800:
                cursor.value = timedelta(seconds=StraDay)
                    
        cursor = sheet.cell(row=1+3,column=13)
        cursor.value = len(totDLav)
        
        totali = calcoloOre(totOLav)
        
        cursor = sheet.cell(row=1+4,column=13)
        cursor.value = len(totNotLav)
        
        cursor = sheet.cell(row=1+5,column=13)
        cursor.value = len(totNoBadge)
        
        cursor = sheet.cell(row=1+6,column=13)
        cursor.value = len(totPossExtra)
        
        cursor = sheet.cell(row=1+7,column=13)
        cursor.value = totali
        
        cursor = sheet.cell(row=1+8,column=13)
        cursor.value = timedelta(seconds=totAnt)
        
        cursor = sheet.cell(row=1+9,column=13)
        cursor.value = timedelta(seconds=totRit)
        
        cursor = sheet.cell(row=1+10,column=13)
        cursor.value = timedelta(seconds=totExtra)
        
        cursor = sheet.cell(row=1+11,column=13)
        cursor.value = timedelta(seconds=totStraDay)
        
        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)
            
        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'fattachment; filename="Ingressi-{dataList[0]}-{ingressi[0][0].title()}.xlsx'
        return response     
           
    else:
        getId = AnaDipendenti.objects.get(user_id=request.user.pk)
        giorniFestivi = so.getHolidays(int(anno))
        festivita = []
        for el in giorniFestivi:
            fes = f'{el.year}-'
            if el.month < 10:
                fes += f'0{el.month}-'
                if el.day < 10:
                    fes += f'0{el.day}'
                else: 
                    fes += f'{el.day}'
            else:
                fes += f'{el.month}-'
                if el.day < 10:
                    fes += f'0{el.day}'
                else: 
                    fes += f'{el.day}'
            festivita.append(fes)
        ingressi = Ingressidip.objects.filter(id_dip_ing=getId.id_dip).values_list("nominativo","entrata","uscita","tipo","id_dip_ing__societa__nome_societa","id_dip_ing__sede__nome_sede","id_dip_ing__area__nome_area","id_dip_ing__tipo_contratto__nome_contratto").order_by('nominativo')
        data = datetime.now().month
        wb = Workbook()
        
        dataList = switch(data)
        sheet = wb.create_sheet(f'T_Ingressi {dataList[0]}')    
        cursor = sheet.cell(row=1,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = "Dipendente"
        
        cursor = sheet.cell(row=1+1,column=1)
        cursor.value = ingressi[0][0]
        
        cursor = sheet.cell(row=1,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 15
        cursor.value = "Società"
        
        cursor = sheet.cell(row=1+1,column=2)
        cursor.value = ingressi[0][4]
        
        cursor = sheet.cell(row=1,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 15
        cursor.value = "Sede"
        
        cursor = sheet.cell(row=1+1,column=3)
        cursor.value = ingressi[0][5]
        
        cursor = sheet.cell(row=1,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 20
        cursor.value = "Area"
        
        cursor = sheet.cell(row=1+1,column=4)
        cursor.value = ingressi[0][6]
        
        cursor = sheet.cell(row=1+3,column=1)
        cursor.font = Font(bold=True)
        cursor.value = "Giorno"
        
        cursor = sheet.cell(row=1+3,column=2)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=1+3,column=3)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=1+3,column=4)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=1+3,column=5)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['E'].width = 20
        cursor.value = "Seconda Entrata"
        
        cursor = sheet.cell(row=1+3,column=6)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['F'].width = 20
        cursor.value = "Seconda Uscita"
        
        cursor = sheet.cell(row=1+3,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=1+3,column=8)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=1+3,column=9)
        cursor.font = Font(bold=True)
        cursor.value = "Extra"
        
        sheet.column_dimensions['J'].width = 15
        cursor = sheet.cell(row=1+3,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Possibili Stra."
        
        cursor = sheet.cell(row=1+3,column=12)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['L'].width = 30
        cursor.value = "Giorni Lavorati"
        
        cursor = sheet.cell(row=1+4,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni di assenza"
        
        cursor = sheet.cell(row=1+5,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni con mancata timbratura"
        
        cursor = sheet.cell(row=1+6,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Giorni con possibili straordinari"
        
        cursor = sheet.cell(row=1+7,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Ore Totali lavorate"
        
        cursor = sheet.cell(row=1+8,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Anticipo (Entrata)"
        
        cursor = sheet.cell(row=1+9,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Ritardo (Entrata)"
        
        cursor = sheet.cell(row=1+10,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Extra (Uscita)"
        
        cursor = sheet.cell(row=1+11,column=12)
        cursor.font = Font(bold=True)
        cursor.value = "Tot. Poss. Stra Mese"

        query = Ingressidip.objects.filter(id_dip_ing=getId.id_dip,giorno__range=[f'{str(today.year)}-{str(dataList[1])}-01',f'{str(anno)}-{str(dataList[1])}-{str(dataList[2])}']).values_list("giorno","entrata","uscita","tipo","tipo_permesso__codicepermesso","id_permesso__da_ora_richiesta","id_permesso__a_ora_richiesta","in_permesso","seconda_entrata","seconda_uscita").order_by("giorno")

        totDLav = []
        totOLav = []
        totELav = []
        totNotLav = []
        totNoBadge = []
        totPossExtra = []
        totAnt = 0
        totRit = 0
        totExtra = 0
        totStraDay = 0
        
        for idx, element in enumerate(query,start=1):
            extraDay = 0
            antiDay = 0
            StraDay = 0
            cursor = sheet.cell(row=idx+4,column=1)
            cursor.font = Font(bold=False)
            cursor.alignment = Alignment(horizontal="right")
            day = str(element[0])
            giornata = element[0]
            entrata = element[1]
            uscita = element[2]
            tipo = element[3]
            codice_permesso = element[4]
            inizio_richiesta = element[5]
            fine_richiesta = element[6]
            in_permesso = element[7]
            seconda_entrata = element[8]
            seconda_uscita = element[9]
            sabDom = giornata.weekday()
            itaDay = f'{day[8:10]}-{day[5:7]}-{day[0:4]}'
            cursor.value = itaDay
            
            if codice_permesso != None and sabDom != 5 and sabDom != 6:
                cursor = sheet.cell(row=idx+4,column=2)
                cursor.alignment = Alignment(horizontal="right")
                cursor.font = Font(bold=True)
                cursor.value = f'{codice_permesso}'
                dayOne = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                dayTwo = ((datetime.strptime(f'1900-01-01 08:00:00','%Y-%m-%d %H:%M:%S')))
                if datetime.today().now().date() > giornata:
                    totOLav.append((dayOne.time(),dayTwo.time()))
                    if sabDom != 5 and sabDom != 6 and entrata == None and uscita == None and in_permesso == 0 and seconda_entrata == None and seconda_uscita == None :
                        totNotLav.append(giornata)
            elif inizio_richiesta != None and fine_richiesta != None:
                cursor = sheet.cell(row=idx+4,column=2)
                cursor.alignment = Alignment(horizontal="right")
                cursor.font = Font(bold=True)
                cursor.value = f'{str(inizio_richiesta)[0:5]} - {str(fine_richiesta)[0:5]}'
            
            if entrata and (sabDom != 5 and sabDom != 6):
                totDLav.append(giornata)
            
            if datetime.today().now().date() > giornata:
                if giornata not in festivita:
                    if uscita == None and (sabDom != 5 and sabDom != 6):
                        totNoBadge.append(giornata)
                    elif entrata == None and (sabDom != 5 and sabDom != 6):
                        totNoBadge.append(giornata)
                
            cursor = sheet.cell(row=idx+4,column=3)
            for el in festivita:
                if el == day and not(entrata):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Festa"
                        break
                elif el == day and entrata:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = entrata
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                elif sabDom == 5 and not(entrata):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Sabato"
                        break
                elif sabDom == 5 and entrata:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = entrata
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                elif sabDom == 6 and not(entrata):
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = "Domenica"
                        break
                elif sabDom == 6 and entrata:
                        cursor.alignment = Alignment(horizontal="right")
                        cursor.font = Font(bold=True,color='FF0000')
                        cursor.value = entrata
                        cursor2 = sheet.cell(row=idx+4,column=4)
                        cursor2.alignment = Alignment(horizontal="right")
                        cursor2.font = Font(bold=True,color='FF0000')
                        break
                else:
                        cursor.value = entrata
                        
            
            cursor = sheet.cell(row=idx+4,column=4)
            cursor.value = uscita
            
            cursor = sheet.cell(row=idx+4,column=5)
            if seconda_entrata:
                cursor.value = seconda_entrata
            
            cursor = sheet.cell(row=idx+4,column=6)
            if seconda_uscita:
                cursor.value = seconda_uscita
            
            cursor = sheet.cell(row=idx+4,column=7)
            cursor.alignment = Alignment(horizontal="right")
            anti = anticipoFunc(entrata)
            if anti and (sabDom != 5 or sabDom != 6):
                antiPre = transDate(anti)
                compare = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                antiDaAgg = antiPre - compare
                antiDay += antiDaAgg.total_seconds()
                if antiDaAgg.total_seconds() > 1800:
                    StraDay += antiDaAgg.total_seconds()
                    totStraDay += antiDaAgg.total_seconds()
                    if giornata not in totPossExtra:
                        totPossExtra.append(giornata)
                totAnt += antiDaAgg.total_seconds()
                cursor.value = antiDaAgg
            else: cursor.value = anti

            cursor = sheet.cell(row=idx+4,column=8)
            cursor.alignment = Alignment(horizontal="right")
            
            rita = ritardoFunc(entrata)
            if rita and (sabDom != 5 or sabDom != 6):
                ritaPre = transDate(rita)
                compare2 = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                ritaDaAgg = ritaPre - compare2
                totRit += ritaDaAgg.total_seconds()
                totRit += 300 
                ritaDaAgg = ritaDaAgg + timedelta(minutes = 5)
                cursor.value = ritaDaAgg
            else: cursor.value = rita
            totOLav.append((entrata,uscita))
            
            secondaEntrata = seconda_entrata
            secondaUscita = seconda_uscita
            
            if secondaEntrata != None and secondaUscita != None:
                primaSeconda = transDate(secondaEntrata)
                comparePrima = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                primSecDaAgg = primaSeconda - comparePrima
                
                secondaSeconda = transDate(secondaUscita)
                compareSeconda = ((datetime.strptime(f'1900-01-01 00:05:00','%Y-%m-%d %H:%M:%S')))
                secoSecDaAgg = secondaSeconda - compareSeconda

                totPrim = primSecDaAgg.total_seconds()
                totSec = secoSecDaAgg.total_seconds()
                totSecBoth = totPrim + totSec
                totOLav.append([seconda_entrata,seconda_uscita])
            
            cursor = sheet.cell(row=idx+4,column=9)
            cursor.alignment = Alignment(horizontal="right")
            extra = extraFunc(uscita)
            if extra and (sabDom != 5 or sabDom != 6):
                extraPre = transDate(extra)
                compare3 = ((datetime.strptime(f'1900-01-01 00:00:00','%Y-%m-%d %H:%M:%S')))
                extraDaAgg = extraPre - compare3
                if extraDaAgg.total_seconds() >= 1800:
                    StraDay += extraDaAgg.total_seconds()
                    totStraDay += extraDaAgg.total_seconds()
                    if giornata not in totPossExtra:
                        totPossExtra.append(giornata)
                elif extraDaAgg.total_seconds() <= 1800:
                    totExtra += extraDaAgg.total_seconds()
                    cursor.value = extra
            totELav.append((entrata,uscita))
            
            cursor = sheet.cell(row=idx+4,column=10)
            cursor.alignment = Alignment(horizontal="right")
            if StraDay > 1800:
                cursor.value = timedelta(seconds=StraDay)
                    
        cursor = sheet.cell(row=1+3,column=13)
        cursor.value = len(totDLav)
        
        totali = calcoloOre(totOLav)
        
        cursor = sheet.cell(row=1+4,column=13)
        cursor.value = len(totNotLav)
        
        cursor = sheet.cell(row=1+5,column=13)
        cursor.value = len(totNoBadge)
        
        cursor = sheet.cell(row=1+6,column=13)
        cursor.value = len(totPossExtra)
        
        cursor = sheet.cell(row=1+7,column=13)
        cursor.value = totali
        
        cursor = sheet.cell(row=1+8,column=13)
        cursor.value = timedelta(seconds=totAnt)
        
        cursor = sheet.cell(row=1+9,column=13)
        cursor.value = timedelta(seconds=totRit)
        
        cursor = sheet.cell(row=1+10,column=13)
        cursor.value = timedelta(seconds=totExtra)
        
        cursor = sheet.cell(row=1+11,column=13)
        cursor.value = timedelta(seconds=totStraDay)
        
        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)
            
        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'fattachment; filename="Ingressi-{dataList[0]}-{ingressi[0][0].title()}.xlsx'
        return response            



@login_required
def downloadDynamicPresenze(request):
    try:
        data =request.POST.get("data")
        data = str(data)
        ingressi = Ingressidip.objects.filter(giorno=data).values_list("id_dip_ing",flat=True)
        wb = Workbook()
        
        sheet = wb.create_sheet(f'T_Ingressi - {data}')
        cell_range = 'A1:L1'
        cursor = sheet.merge_cells(cell_range)
        cursor = sheet.cell(row=1,column=1)           
        cursor.alignment = Alignment(horizontal='center')
        cursor.value = f"Presenze del giorno {str(data)}"
        
        cursor = sheet.cell(row=2,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = "Dipendente"
        
        cursor = sheet.cell(row=2,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 25
        cursor.value = "Società"
  
        cursor = sheet.cell(row=2,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 20
        cursor.value = "Sede"
        
        cursor = sheet.cell(row=2,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 25
        cursor.value = "Area"
        
        cursor = sheet.cell(row=2,column=5)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=2,column=6)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=2,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=2,column=8)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['H'].width = 20
        cursor.value = "Seconda Entrata"
        
        cursor = sheet.cell(row=2,column=9)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['I'].width = 20
        cursor.value = "Seconda Uscita"
        
        cursor = sheet.cell(row=2,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=2,column=11)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=2,column=12)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['L'].width = 20
        cursor.value = "Poss. Straordinari"

        for index,el in enumerate(ingressi,start=1):
            try:
                dipendente = Ingressidip.objects.get(giorno=data,id_dip_ing=ingressi[index])
            except:
                pass 
           
            try:
                if dipendente.nominativo:
                    cursor = sheet.cell(row=2+index,column=1)
                    cursor.value = str(dipendente.nominativo)
            except:
                pass  
             
            try:
                if dipendente.id_dip_ing.societa:
                    cursor = sheet.cell(row=2+index,column=2)
                    cursor.value = str(dipendente.id_dip_ing.societa).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.sede:
                    cursor = sheet.cell(row=2+index,column=3)
                    cursor.value = str(dipendente.id_dip_ing.sede).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.area:
                    cursor = sheet.cell(row=2+index,column=4)
                    cursor.value = str(dipendente.id_dip_ing.area.nome_area).title()
            except:
                pass
            try:
                if dipendente.in_permesso:
                    cursor = sheet.cell(row=2+index,column=5)
                    if dipendente.id_permesso:
                        if dipendente.id_permesso.codicepermesso:
                            cursor.value = str(dipendente.dipendente.id_permesso.codicepermesso)
                        else: cursor.value = "Permesso Orario"
                else: 
                    cursor = sheet.cell(row=2+index,column=5)
                    cursor.value = ""
            except:
                pass
            try:
                if dipendente.entrata:
                    cursor = sheet.cell(row=2+index,column=6)
                    cursor.value = str(dipendente.entrata)
            except:
                pass
            try:
                if dipendente.uscita:
                    cursor = sheet.cell(row=2+index,column=7)
                    cursor.value = str(dipendente.uscita)
            except:
                pass
            try:
                if dipendente.seconda_entrata:
                    cursor = sheet.cell(row=2+index,column=8)
                    cursor.value = str(dipendente.seconda_entrata)
            except:
                pass
            try:
                if dipendente.seconda_uscita:
                    cursor = sheet.cell(row=2+index,column=9)
                    cursor.value = str(dipendente.seconda_uscita)
            except:
                pass
            try:
                if dipendente.anticipo:
                    cursor = sheet.cell(row=2+index,column=10)
                    cursor.value = str(dipendente.anticipo)
            except:
                pass
            try:
                if dipendente.ritardo:
                    cursor = sheet.cell(row=2+index,column=11)
                    cursor.value = str(dipendente.ritardo)
            except:
                pass
            try:
                if dipendente.straordinario:
                    cursor = sheet.cell(row=2+index,column=12)
                    cursor.value = str(dipendente.straordinario)
            except:
                pass
        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)
            
        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'fattachment; filename="Ingressi-{data}.xlsx'
        return response
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required
def downloadAllPresenzeDipendente(request):
    try:
        data =request.POST.get("data")
        data = str(data)
        ingressi = Ingressidip.objects.filter(giorno=data).values_list("id_dip_ing",flat=True)
        wb = Workbook()
        
        sheet = wb.create_sheet(f'T_Ingressi - {data}')
        cell_range = 'A1:L1'
        cursor = sheet.merge_cells(cell_range)
        cursor = sheet.cell(row=1,column=1)           
        cursor.alignment = Alignment(horizontal='center')
        cursor.value = f"Presenze del giorno {str(data)}"
        
        cursor = sheet.cell(row=2,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = "Dipendente"
        
        cursor = sheet.cell(row=2,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 25
        cursor.value = "Società"
  
        cursor = sheet.cell(row=2,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 20
        cursor.value = "Sede"
        
        cursor = sheet.cell(row=2,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 25
        cursor.value = "Area"
        
        cursor = sheet.cell(row=2,column=5)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=2,column=6)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=2,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=2,column=8)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['H'].width = 20
        cursor.value = "Seconda Entrata"
        
        cursor = sheet.cell(row=2,column=9)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['I'].width = 20
        cursor.value = "Seconda Uscita"
        
        cursor = sheet.cell(row=2,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=2,column=11)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=2,column=12)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['L'].width = 20
        cursor.value = "Poss. Straordinari"

        for index,el in enumerate(ingressi,start=1):
            try:
                dipendente = Ingressidip.objects.get(giorno=data,id_dip_ing=ingressi[index])
            except:
                pass 
           
            try:
                if dipendente.nominativo:
                    cursor = sheet.cell(row=2+index,column=1)
                    cursor.value = str(dipendente.nominativo)
            except:
                pass  
             
            try:
                if dipendente.id_dip_ing.societa:
                    cursor = sheet.cell(row=2+index,column=2)
                    cursor.value = str(dipendente.id_dip_ing.societa).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.sede:
                    cursor = sheet.cell(row=2+index,column=3)
                    cursor.value = str(dipendente.id_dip_ing.sede).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.area:
                    cursor = sheet.cell(row=2+index,column=4)
                    cursor.value = str(dipendente.id_dip_ing.area.nome_area).title()
            except:
                pass
            try:
                if dipendente.in_permesso:
                    cursor = sheet.cell(row=2+index,column=5)
                    if dipendente.id_permesso:
                        if dipendente.id_permesso.codicepermesso:
                            cursor.value = str(dipendente.dipendente.id_permesso.codicepermesso)
                        else: cursor.value = "Permesso Orario"
                else: 
                    cursor = sheet.cell(row=2+index,column=5)
                    cursor.value = ""
            except:
                pass
            try:
                if dipendente.entrata:
                    cursor = sheet.cell(row=2+index,column=6)
                    cursor.value = str(dipendente.entrata)
            except:
                pass
            try:
                if dipendente.uscita:
                    cursor = sheet.cell(row=2+index,column=7)
                    cursor.value = str(dipendente.uscita)
            except:
                pass
            try:
                if dipendente.seconda_entrata:
                    cursor = sheet.cell(row=2+index,column=8)
                    cursor.value = str(dipendente.seconda_entrata)
            except:
                pass
            try:
                if dipendente.seconda_uscita:
                    cursor = sheet.cell(row=2+index,column=9)
                    cursor.value = str(dipendente.seconda_uscita)
            except:
                pass
            try:
                if dipendente.anticipo:
                    cursor = sheet.cell(row=2+index,column=10)
                    cursor.value = str(dipendente.anticipo)
            except:
                pass
            try:
                if dipendente.ritardo:
                    cursor = sheet.cell(row=2+index,column=11)
                    cursor.value = str(dipendente.ritardo)
            except:
                pass
            try:
                if dipendente.straordinario:
                    cursor = sheet.cell(row=2+index,column=12)
                    cursor.value = str(dipendente.straordinario)
            except:
                pass
        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)
            
        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'fattachment; filename="Ingressi-{data}.xlsx'
        return response
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required
def downloadDynamicPresenze(request):
    try:
        data =request.POST.get("data")
        data = str(data)
        ingressi = Ingressidip.objects.filter(giorno=data).values_list("id_dip_ing",flat=True)
        wb = Workbook()
        
        sheet = wb.create_sheet(f'T_Ingressi - {data}')
        cell_range = 'A1:L1'
        cursor = sheet.merge_cells(cell_range)
        cursor = sheet.cell(row=1,column=1)           
        cursor.alignment = Alignment(horizontal='center')
        cursor.value = f"Presenze del giorno {str(data)}"
        
        cursor = sheet.cell(row=2,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = "Dipendente"
        
        cursor = sheet.cell(row=2,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 25
        cursor.value = "Società"
  
        cursor = sheet.cell(row=2,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 20
        cursor.value = "Sede"
        
        cursor = sheet.cell(row=2,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 25
        cursor.value = "Area"
        
        cursor = sheet.cell(row=2,column=5)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=2,column=6)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=2,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=2,column=8)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['H'].width = 20
        cursor.value = "Seconda Entrata"
        
        cursor = sheet.cell(row=2,column=9)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['I'].width = 20
        cursor.value = "Seconda Uscita"
        
        cursor = sheet.cell(row=2,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=2,column=11)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=2,column=12)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['L'].width = 20
        cursor.value = "Poss. Straordinari"

        for index,el in enumerate(ingressi,start=1):
            try:
                dipendente = Ingressidip.objects.get(giorno=data,id_dip_ing=ingressi[index])
            except:
                pass 
           
            try:
                if dipendente.nominativo:
                    cursor = sheet.cell(row=2+index,column=1)
                    cursor.value = str(dipendente.nominativo)
            except:
                pass  
             
            try:
                if dipendente.id_dip_ing.societa:
                    cursor = sheet.cell(row=2+index,column=2)
                    cursor.value = str(dipendente.id_dip_ing.societa).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.sede:
                    cursor = sheet.cell(row=2+index,column=3)
                    cursor.value = str(dipendente.id_dip_ing.sede).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.area:
                    cursor = sheet.cell(row=2+index,column=4)
                    cursor.value = str(dipendente.id_dip_ing.area.nome_area).title()
            except:
                pass
            try:
                if dipendente.in_permesso:
                    cursor = sheet.cell(row=2+index,column=5)
                    if dipendente.id_permesso:
                        if dipendente.id_permesso.codicepermesso:
                            cursor.value = str(dipendente.dipendente.id_permesso.codicepermesso)
                        else: cursor.value = "Permesso Orario"
                else: 
                    cursor = sheet.cell(row=2+index,column=5)
                    cursor.value = ""
            except:
                pass
            try:
                if dipendente.entrata:
                    cursor = sheet.cell(row=2+index,column=6)
                    cursor.value = str(dipendente.entrata)
            except:
                pass
            try:
                if dipendente.uscita:
                    cursor = sheet.cell(row=2+index,column=7)
                    cursor.value = str(dipendente.uscita)
            except:
                pass
            try:
                if dipendente.seconda_entrata:
                    cursor = sheet.cell(row=2+index,column=8)
                    cursor.value = str(dipendente.seconda_entrata)
            except:
                pass
            try:
                if dipendente.seconda_uscita:
                    cursor = sheet.cell(row=2+index,column=9)
                    cursor.value = str(dipendente.seconda_uscita)
            except:
                pass
            try:
                if dipendente.anticipo:
                    cursor = sheet.cell(row=2+index,column=10)
                    cursor.value = str(dipendente.anticipo)
            except:
                pass
            try:
                if dipendente.ritardo:
                    cursor = sheet.cell(row=2+index,column=11)
                    cursor.value = str(dipendente.ritardo)
            except:
                pass
            try:
                if dipendente.straordinario:
                    cursor = sheet.cell(row=2+index,column=12)
                    cursor.value = str(dipendente.straordinario)
            except:
                pass
        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)
            
        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'fattachment; filename="Ingressi-{data}.xlsx'
        return response
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required
def downloadAllPresenzeDipendente(request):
    try:
        data =request.POST.get("data")
        data = str(data)
        ingressi = Ingressidip.objects.filter(giorno=data).values_list("id_dip_ing",flat=True)
        wb = Workbook()
        
        sheet = wb.create_sheet(f'T_Ingressi - {data}')
        cell_range = 'A1:L1'
        cursor = sheet.merge_cells(cell_range)
        cursor = sheet.cell(row=1,column=1)           
        cursor.alignment = Alignment(horizontal='center')
        cursor.value = f"Presenze del giorno {str(data)}"
        
        cursor = sheet.cell(row=2,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = "Dipendente"
        
        cursor = sheet.cell(row=2,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 25
        cursor.value = "Società"
  
        cursor = sheet.cell(row=2,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 20
        cursor.value = "Sede"
        
        cursor = sheet.cell(row=2,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 25
        cursor.value = "Area"
        
        cursor = sheet.cell(row=2,column=5)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=2,column=6)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=2,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=2,column=8)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['H'].width = 20
        cursor.value = "Seconda Entrata"
        
        cursor = sheet.cell(row=2,column=9)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['I'].width = 20
        cursor.value = "Seconda Uscita"
        
        cursor = sheet.cell(row=2,column=10)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=2,column=11)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=2,column=12)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['L'].width = 20
        cursor.value = "Poss. Straordinari"

        for index,el in enumerate(ingressi,start=1):
            try:
                dipendente = Ingressidip.objects.get(giorno=data,id_dip_ing=ingressi[index])
            except:
                pass 
           
            try:
                if dipendente.nominativo:
                    cursor = sheet.cell(row=2+index,column=1)
                    cursor.value = str(dipendente.nominativo)
            except:
                pass  
             
            try:
                if dipendente.id_dip_ing.societa:
                    cursor = sheet.cell(row=2+index,column=2)
                    cursor.value = str(dipendente.id_dip_ing.societa).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.sede:
                    cursor = sheet.cell(row=2+index,column=3)
                    cursor.value = str(dipendente.id_dip_ing.sede).title()
            except:
                pass
            try:
                if dipendente.id_dip_ing.area:
                    cursor = sheet.cell(row=2+index,column=4)
                    cursor.value = str(dipendente.id_dip_ing.area.nome_area).title()
            except:
                pass
            try:
                if dipendente.in_permesso:
                    cursor = sheet.cell(row=2+index,column=5)
                    if dipendente.id_permesso:
                        if dipendente.id_permesso.codicepermesso:
                            cursor.value = str(dipendente.dipendente.id_permesso.codicepermesso)
                        else: cursor.value = "Permesso Orario"
                else: 
                    cursor = sheet.cell(row=2+index,column=5)
                    cursor.value = ""
            except:
                pass
            try:
                if dipendente.entrata:
                    cursor = sheet.cell(row=2+index,column=6)
                    cursor.value = str(dipendente.entrata)
            except:
                pass
            try:
                if dipendente.uscita:
                    cursor = sheet.cell(row=2+index,column=7)
                    cursor.value = str(dipendente.uscita)
            except:
                pass
            try:
                if dipendente.seconda_entrata:
                    cursor = sheet.cell(row=2+index,column=8)
                    cursor.value = str(dipendente.seconda_entrata)
            except:
                pass
            try:
                if dipendente.seconda_uscita:
                    cursor = sheet.cell(row=2+index,column=9)
                    cursor.value = str(dipendente.seconda_uscita)
            except:
                pass
            try:
                if dipendente.anticipo:
                    cursor = sheet.cell(row=2+index,column=10)
                    cursor.value = str(dipendente.anticipo)
            except:
                pass
            try:
                if dipendente.ritardo:
                    cursor = sheet.cell(row=2+index,column=11)
                    cursor.value = str(dipendente.ritardo)
            except:
                pass
            try:
                if dipendente.straordinario:
                    cursor = sheet.cell(row=2+index,column=12)
                    cursor.value = str(dipendente.straordinario)
            except:
                pass
        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)
            
        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'fattachment; filename="Ingressi-{data}.xlsx'
        return response
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required
@csrf_exempt
def downloadAllPresenzePerDipendente(request):
    try:
        data = datetime.now().date()
        dipendente = request.POST.get("dipendente")
        ingressi = Ingressidip.objects.filter(giorno=data,id_dip_ing=dipendente)
        dip = AnaDipendenti.objects.get(id_dip=dipendente)
        wb = Workbook()
        
        sheet = wb.create_sheet(f'T_Ingressi - {data}')
        cell_range = 'A1:L1'
        cursor = sheet.merge_cells(cell_range)
        cursor = sheet.cell(row=1,column=1)           
        cursor.alignment = Alignment(horizontal='center')
        cursor.value = f"Presenze del dipendente {str(dip.nominativo)}"
        
        cursor = sheet.cell(row=2,column=1)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['A'].width = 30
        cursor.value = str(dipendente.nominativo)
        
        cursor = sheet.cell(row=2,column=2)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['B'].width = 25
        cursor.value = str(dipendente.id_dip_ing.societa).title()
  
        cursor = sheet.cell(row=2,column=3)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 20
        cursor.value = str(dipendente.id_dip_ing.sede).title()
        
        cursor = sheet.cell(row=2,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 25
        cursor.value = str(dipendente.id_dip_ing.area.nome_area).title()
        
        cursor = sheet.cell(row=3,column=1)
        cursor.font = Font(bold=True)
        cursor.value = "Permesso"
        
        cursor = sheet.cell(row=3,column=2)
        cursor.font = Font(bold=True)
        cursor.value = "Entrata"
        
        cursor = sheet.cell(row=3,column=3)
        cursor.font = Font(bold=True)
        cursor.value = "Uscita"
        
        cursor = sheet.cell(row=3,column=4)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['C'].width = 20
        cursor.value = "Seconda Entrata"
        
        cursor = sheet.cell(row=3,column=5)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['D'].width = 20
        cursor.value = "Seconda Uscita"
        
        cursor = sheet.cell(row=3,column=6)
        cursor.font = Font(bold=True)
        cursor.value = "Anticipo"
        
        cursor = sheet.cell(row=3,column=7)
        cursor.font = Font(bold=True)
        cursor.value = "Ritardo"
        
        cursor = sheet.cell(row=3,column=8)
        cursor.font = Font(bold=True)
        sheet.column_dimensions['G'].width = 20
        cursor.value = "Poss. Straordinari"
        
        for el,index in enumerate(ingressi,start=1):
            try:
                if dipendente.in_permesso:
                    cursor = sheet.cell(row=4+index,column=1)
                    if dipendente.id_permesso:
                        if dipendente.id_permesso.codicepermesso:
                            cursor.value = str(dipendente.dipendente.id_permesso.codicepermesso)
                        else: cursor.value = "Permesso Orario"
                else: 
                    cursor = sheet.cell(row=4+index,column=1)
                    cursor.value = ""
            except:
                pass
            try:
                if dipendente.entrata:
                    cursor = sheet.cell(row=4+index,column=2)
                    cursor.value = str(dipendente.entrata)
            except:
                pass
            try:
                if dipendente.uscita:
                    cursor = sheet.cell(row=4+index,column=3)
                    cursor.value = str(dipendente.uscita)
            except:
                pass
            try:
                if dipendente.seconda_entrata:
                    cursor = sheet.cell(row=4+index,column=4)
                    cursor.value = str(dipendente.seconda_entrata)
            except:
                pass
            try:
                if dipendente.seconda_uscita:
                    cursor = sheet.cell(row=4+index,column=5)
                    cursor.value = str(dipendente.seconda_uscita)
            except:
                pass
            try:
                if dipendente.anticipo:
                    cursor = sheet.cell(row=4+index,column=6)
                    cursor.value = str(dipendente.anticipo)
            except:
                pass
            try:
                if dipendente.ritardo:
                    cursor = sheet.cell(row=4+index,column=7)
                    cursor.value = str(dipendente.ritardo)
            except:
                pass
            try:
                if dipendente.straordinario:
                    cursor = sheet.cell(row=4+index,column=8)
                    cursor.value = str(dipendente.straordinario)
            except:
                pass
        try:
            del wb['Sheet']
        except:
            pass
        output = BytesIO()
        wb.save(output)
        output.seek(0)
            
        response = HttpResponse(output.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'fattachment; filename="Ingressi-{dip.nominativo}.xlsx'
        return response
    except: 
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))